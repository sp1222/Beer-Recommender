'''Beer Classifier and Recommender

Use defined features of beer objects from data collection
with kNN algorithm to classify and recommend beer.

Displays the results with graphical representation


Imports
----------
collections
    Counter
math
    pi
openpyxl
    Workbook
    load_workbook
sklearn.preprocessing
    normalize
BeerClass
matplotlib
numpy
os
random
re
time


'''

from collections import Counter
from math import pi
from openpyxl import Workbook
from openpyxl import load_workbook
from sklearn.preprocessing import normalize
import BeerClass
import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import os
import random
import re
import time

# CONSTANTS
BEER_FEATURES_START_ROW = 50

# the starting row of category information
CATEGORY_FEATURES_START_ROW = 15

# number of features in a beer
MAX_NUMBER_OF_FEATURES = 11

# File Names:
BEER_ALL_INFO = 'Beer_All_Info\\'
FILE_DIRECTORY = 'D:\\Python Projects\\Beer Recommender Project\\'
USER_INPUT_DIRECTORY = 'User Input\\'


#********************************************************************************************************************************
# this function loads the basic information for all 5700 beer for machine learning application
# with the option to load user beer input from file as well.
def loadBeerInformation(getUser = False, getCategoryDictionary = False, normalizeFeatures = False):
    ''' Returns a list of beer objects loaded from xlsx style/category files

    Parameters
    ----------
    getUser : bool
        determines if function returns static user input data
        defaults to False
        
    getCategoryDictionary : bool
        determines if function returns category dictionary
        defaults to False

    normalizeFeatures : bool
        determines if function normalizes all of the BeerClass\'s feature matrix
        defaults to False
        UNDER CONSTRUCTION

    Returns
    ----------
    beerList : list
        list of BeerClass objects
        beerList will always be returned

    userBeerList : list
        list of BeerClass objects as static user input
        returns if getUser == True

    userCV : list
        list of BeerClass objects for cross validation
        returns if getUser == True

    categoryDictionary : dict
        dictionary of key-value pairs "categoryKey: categoryName"
        dictionary[ float(categoryKey) ] = str(categoryName)
        returns if getUser == True and getCategoryDictionary == True


    '''
    
    beerList = []
    normalized = np.empty(MAX_NUMBER_OF_FEATURES)
    if getCategoryDictionary == True:
        categoryDictionary = {}

    # first we need to see if the name of the categories we want to collect data from exist as files
    fileList = os.listdir(FILE_DIRECTORY + BEER_ALL_INFO)
    print('Gathering File Names..')
    fileName = BEER_ALL_INFO
    for file in fileList:
        wb = load_workbook(fileName + file)
        if getCategoryDictionary == True:
            index = 0
            wb.active = index
            sheet = wb.active
            categoryDictionary[float(sheet.cell(row = 2, column = 2).value)] = str(sheet.cell(row = 1, column = 2).value)
        
        index = 1
        print('Getting beer information from ' + file)
        while index < len(wb.worksheets):
            beer = BeerClass.Beer()            
            wb.active = index
            sheet = wb.active
            currentRow = 1
            currentColumn = 2
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerName(re.sub('\n','',currentCell.value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerKey(int(currentCell.value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerStyle(re.sub('\n','',currentCell.value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerCategoryKey(int(currentCell.value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerBrewery(re.sub('\n','',currentCell.value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerAverageRating(float(currentCell.value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerABV(float(currentCell.value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerMinIBU(int(currentCell.value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerMaxIBU(int(currentCell.value))

            features = []
            currentRow = BEER_FEATURES_START_ROW
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            while currentCell.value != '' and currentCell.value != None:
                features.append(int(currentCell.value))
                currentRow += 1
                currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerFeaturesMatrix(features)   
            beerList.append(beer)
            index += 1

#### LEFT OFF HERE WITH NORMALIZING DATA
    if normalizeFeatures == True:
        for b in beerList:
            normalized = np.vstack((normalized, np.array(b.getBeerFeaturesMatrix)))        
        # we need to delete the first row that came from np.empty before normalizing.
        beerFeatures = np.delete(beerFeatures, 0, 0)
        normalized = normalize(normalized, axis = 0, norm = 'l1')
        for i in range(beerList):
            b.setBeerFeaturesMatrix(normalized[i,:].tolist())
####

    if getUser == True:            
        wb = load_workbook(USER_INPUT_DIRECTORY + 'User Input.xlsx')
        index = 1
        wb.active = index
        sheet = wb.active
        userBeerList = []
        userCV = []

        while index < len(wb.worksheets):        
            beer = BeerClass.Beer()
            beerCV = BeerClass.Beer()
            wb.active = index
            sheet = wb.active
            currentRow = 1
            currentColumn = 2
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerName(re.sub('\n','',currentCell.value))
            beerCV.setBeerName(re.sub('\n','',currentCell.value) + ' :: Cross Validation')
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerKey(int(currentCell.value))
            beerCV.setBeerKey(int(currentCell.value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerStyle(re.sub('\n','',currentCell.value))
            beerCV.setBeerStyle(re.sub('\n','',currentCell.value) + ' :: Cross Validation')
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerCategoryKey(int(currentCell.value))
            beerCV.setBeerCategoryKey(int(currentCell.value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerBrewery(re.sub('\n','',currentCell.value))
            beerCV.setBeerBrewery(re.sub('\n','',currentCell.value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerAverageRating(float(currentCell.value))
            beerCV.setBeerAverageRating(float(currentCell.value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerABV(float(currentCell.value))
            beerCV.setBeerABV(float(currentCell.value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerMinIBU(int(currentCell.value))
            beerCV.setBeerMinIBU(int(currentCell.value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerMaxIBU(int(currentCell.value))
            beerCV.setBeerMaxIBU(int(currentCell.value))

            features = []
            featuresCV = []
            currentRow = BEER_FEATURES_START_ROW
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            while currentCell.value != '' and currentCell.value != None:
                features.append(int(currentCell.value))
                featuresCV.append(int(currentCell.value + 1))
                currentRow += 1
                currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerFeaturesMatrix(features)
            beerCV.setBeerFeaturesMatrix(features)
            userBeerList.append(beer)
            userCV.append(beerCV)
            index += 1
        if getCategoryDictionary == True:
            return beerList, userBeerList, userCV, categoryDictionary
        else:
            return beerList, userBeerList, userCV
    return beerList

###*******************************************************************************************************************************
# Functions to run K Nearest Neighbor

# data will be a dictionary of beer, with beer keys being key and values being features
# predict is the user's chosen beer to compare with
# k is the number of neighbors we will be looking for.


def __calculateEuclideanDistance(data, predict):
    ''' Calculate Euclidean Distance using beer features matrix
    
    Parameters
    ----------
    data : list
        list of a BeerClass\'s features matrix

    predict : list
        list of a BeerClass\'s features matrix

    Returns
    ----------
    euclideanDistance : float
        the euclidean distance between two sets of beer features matrices    

    '''
    euclideanDistance = np.linalg.norm(np.array(data) - np.array(predict))

    return euclideanDistance


#*****************************************************************************************************************************
# calculates all neighbors from data in dictionaries

def calculateAllNN(aBeer, uInput):
    ''' Calculate the nearest neighbors between all beer objects in dataset and all user input
    
    Parameters
    ----------
    aBeer : dict
        dictionary of all BeerClass objects in dataset
        key-values beerKey : features matrix

    uInput : dict
        dictionary of user input BeerClass objects

    Returns
    ----------
    results : dict
        dictionary of user input keys as keys and all BeerClass keys
        and all BeerClass keys with their distances from user input as values
        the values to user input keys are sorted by distance in ascending order
        layout below
        
        result = 
        {
            userInputKey1: { {allBeerKey1: distance1}, {allBeerKey2, distance2}, ... }
            userInputKey2: { {allBeerKey1: distance1}, {allBeerKey2, distance2}, ... }
                 ....
        }

    '''
    
    results = {}    # dictionary of dictionaries to be returned, set up as each user beer key as the key,
                    # and the value being a dictionary of each beer in all beer with the all beer key as the key and the all beer distance as the value
                    # sorted by distances in ascending order

    # {
    #  userInputKey1: {{allBeerKey1: distance1}, {allBeerKey2, distance2}}
    #  userInputKey2: {{allBeerKey1: distance1}, {allBeerKey2, distance2}}
    #     ....
    # }

    # iterate through each user input of enjoyed beers
    for ukey, uval in uInput.items():
        distances = []

        # for each of the 5600 beers in our data, determine the distance between the current user input
        for bkey, bval in aBeer.items():
            distances.append([__calculateEuclideanDistance(bval, uval), bkey, bval])
            
        # sort by shortest distance to the current user input
        distances = sorted(distances)

        ukeyValues = {}
        for dist in distances:
            ukeyValues.update({dist[1]: dist[0]})
        results.update({ukey: ukeyValues})
    return results


#**********************************************************************************************************************************
# functions to print recommendations based on each userPredict beer.

def printBNearestNeighbors(data, aBeer, k, userInput):
    ''' Print recommendations based on the results of kNN algorithm
    Also prints graphical representation of the results
    
    Parameters
    ----------
    data : dict
        dictionary obtained from calculateAllNN(data, predict)

        data = 
        {
            userInputKey1: { {allBeerKey1: distance1}, {allBeerKey2, distance2}, ... }
            userInputKey2: { {allBeerKey1: distance1}, {allBeerKey2, distance2}, ... }
                 ....
        }
        
    aBeer : dict
        dictionary of key-values beerKey : features matrix
    
    k : int
        k-nearest neighbors

    userInput :
        dictionary of key-values userInputKey : features matrix    

    '''
    
    for dKey, dVal in data.items():
        knn = {}    # a dictionary of dictionaries, with the user input being the first entry of knn dictionary.
        currentUserInputBeer = BeerClass.Beer()
        for userIn in userInput:
            if userIn.getBeerKey() == dKey:
                currentUserInputBeer = userIn
                break
        print('*************************************************************')
        print('\nRecommendations based on ' + currentUserInputBeer.getBeerName())
        print('Style:        ' + currentUserInputBeer.getBeerStyle())
        print(currentUserInputBeer.getBeerFeaturesMatrix())
        # match the key collected from allNN result to the list of beers that are in the allBeer list and print the data.
        counter = 0
        atK = False
        knn.update({currentUserInputBeer.getBeerName(): currentUserInputBeer.getBeerFeaturesMatrix()})
        for bkey, bvalues in dVal.items():
            for beer in aBeer:
                if bkey == beer.getBeerKey():
                    print('\nName:         ' + str(beer.getBeerName()))
                    print('Style:        ' + str(beer.getBeerStyle()))
                    print('Feature Dist: ' + str(bvalues))
                    print(beer.getBeerFeaturesMatrix())
                    knn.update({beer.getBeerName(): beer.getBeerFeaturesMatrix()})
                    counter += 1
                    if counter == k:
                        atK = True
                    break
            if atK == True:
                break
        __graphRecommendations(knn)
#    return knn


#******************************************************************************************************************
# Graphing the user input features and the b nearest neighbors recommended.

def __graphRecommendations(data):
    fig = plt.figure(figsize = (12, 12))
    ax = plt.subplot(polar = 'True')    
    featureLabels = ['Astringency', 'Body', 'Alcoholic', 'Bitter', 'Sweet', 'Sour', 'Salty', 'Fruity', 'Hoppy', 'Spice', 'Malty']   
    color = ['red', 'green', 'blue', 'yellow', 'cyan', 'magenta', 'brown', 'coral', 'darkgreen', 'gold', 'fuchsia', 'lightblue', 'maroon', 'teal', 'violet']
    currentColor = -1
    dataNames = []
    title = list(data.keys())[0]
    # here we unpack our data.
    for name, features in data.items():
        dataNames.append(name)      # for legend chart
        # next we need to caluclate the angles for each feature
        angles = [n / float(MAX_NUMBER_OF_FEATURES) * 2 * pi for n in range(MAX_NUMBER_OF_FEATURES)]
        try:
            # we need to complete the polygon by adding the starting feature at the end.
            features += features[:1]
            # and we need to repate the first value to complete the circle
            angles += angles[:1]
        except:
            print('an error occurred in graphing')
            continue
        try:
            currentColor += 1
            plt.polar(angles, features, marker = '.', color = color[currentColor], label = name)
            
        except:
            print('something happened in plt.polar')
            continue

        plt.xticks(angles[:-1], featureLabels)
    
    ax.set_rlabel_position(0)
    plt.yticks([25, 50, 75, 100, 125, 150], color = 'grey', size = 10)
    plt.ylim(0, 200)
    plt.title(title)
    legend = ax.legend(loc = 'upper right')
    fig.savefig('recommendations for ' + title + '.png')
    plt.show()


#******************************************************************************************************************
# function to classify the user input's beer style using KNN
# this function returns KNN distances and classification label.

def classifyNewBeerUsingBeerObjects(aBeer, nBeer, k):
    ''' Function that uses kNN to classify a new beer based on optimized k values.

    Parameters
    ----------
    aBeer : list
        list of BeerClass objects in dataset
        

    nBeer : BeerClass
        BeerClass object to be classified

    k : int
        k value for kNN algorithm, used to return the k closest BeerClass objects

    Returns
    ----------
    distancesKNN : list
        list of k-nearest BeerClass objects
        sorted in ascending order by distance from new BeerClass object

    mostLabelsInDistancesKNN : int
        integer value representing the most category labels found in k-nearest BeerClass objects

    '''

    distancesAll = []
    for beer in aBeer:
    # distancesAll = [euclideanDistance, beerKey, beerStyleKey]
        distancesAll.append([__calculateEuclideanDistance(beer.getBeerFeaturesMatrix(), nBeer.getBeerFeaturesMatrix()), beer.getBeerKey(), beer.getBeerStyle()])
    distancesSorted = sorted(distancesAll)   # sort the distances by distance from shortest to longest
    distancesKNN = distancesSorted[:k]
    distancesKNN = np.array(distancesKNN)
    mostLabelsInDistancesKNN = Counter(distancesKNN[:,2]).most_common(1)[0][0]
    
    return distancesKNN, mostLabelsInDistancesKNN


#*********************************************************************************************************************
# Data Setup for KNN algorithms:
# Set up our numpy arrays for KNN algorithms
# beerFeatures = [x1, x2, x3, x4, x5, x6, x7, x8, x9, x10, x11]
# beerStyle = [y]
# where xN is the feature from the features matrix

def dataSetup(listOfBeerIn, yAvailable = False, normalizeFeatures = False):
    ''' build numpy arrays from a list of beer objects

    Parameters
    ----------
    listOfBeerIn : list
        list of beer objects
        
    yAvailable : bool
        determines if beer object category loaded into numpy arrays
        defaults to False

    normalizeFeatures : bool
        determines if the numpy array values need to be normalized
        defaults to False
        UNDER CONSTRUCTION

    Returns
    ----------
    beerFeatures : numpy array
        All BeerClass object's features matrix converted to numpy array as x values

    beerStyle : numpy array
        all BeerClass object's categories converted to numpy array as y values        

    '''

    beerFeatures = np.empty(MAX_NUMBER_OF_FEATURES)
    beerStyle = np.empty(1)     # if yAvailable == True.
    
    for beer in listOfBeerIn:
        xrow = np.array(beer.getBeerFeaturesMatrix())
        beerFeatures = np.vstack((beerFeatures, xrow))            
        if yAvailable == True:
            yrow = np.array(beer.getBeerCategoryKey())
            beerStyle = np.vstack((beerStyle, yrow))

    # we need to delete the first row that came from np.empty.
    beerFeatures = np.delete(beerFeatures, 0, 0)
    if yAvailable == True:
        beerStyle = np.delete(beerStyle, 0, 0)

    # normalize features.
    if normalizeFeatures == True:
        beerFeatures = normalize(beerFeatures, axis = 0, norm = 'l1')
        
        
    if yAvailable == True:
        return beerFeatures, beerStyle
    return beerFeatures
        

#*******************************************************************************************************************
# classify beer using a list of type [array(beer.getBeerFeaturesMatrix()), style key]
# this will hopefully save some time.

def getSortedDistancesUsingNumpy(aBeer, nBeer, style):
    ''' Use numpy arrays created in dataSetup to calculate Euclidean distances of beer objects

    Parameters
    ----------
    aBeer : numpy array
        dataset of all BeerClass object\'s features matrix in numpy format
        
    nBeer : BeerClass object
        new BeerClass object\'s features matrix in numpy format
        
    style : numpy array
        dataset of all BeerClass object\' categorys keys in numpy format

    Returns
    ----------
    distancesSorted : list
        BeerClass objects sorted by distance from new BeerClass object in ascending order

     '''


    # calculate euclidean distance between new beer and each beer in all beer.
    euclideanDistance = np.array([ np.linalg.norm(a - nBeer) for a in aBeer])
    euclideanDistance = euclideanDistance.reshape((euclideanDistance.shape[0], 1))
    # aftcalculating euclidean distance, we append the corresponding styles to the array.
    distancesAll = np.append(euclideanDistance, style, axis = 1)
    distancesSorted = distancesAll[distancesAll[:,0].argsort()]   # sort the distances by distance from shortest to longest
    return distancesSorted
    

#********************************************************************************************************************
# Optimizing K

def optimizeKUsingNumpy(data, style, maxTrainingLabelsCount):
    ''' Use test BeerClass set formatted as numpy arrays created in dataSetup to find optimized values of k for kNN algorithm

    Parameters
    ----------
    data : numpy array
        dataset of all BeerClass object\'s features matrix in numpy format
        
    style : numpy array
        dataset of all BeerClass object\' categorys keys in numpy format

    maxTrainingLabelsCount : int
        max number of BeerClass objects from each category to use for training (80%)

    Returns
    ----------
    K : list
        list of ideal values of k to use for classification, returns only those that are 100% accurate

     '''

    highestAccuracyAtK = [0] * len(data)
    accuracyLimit = .8
    K = []
    for currentData, currentStyle in zip(data, style):
        allSortedDistances = getSortedDistancesUsingNumpy(data, currentData, style)[1:]
        currentK = []        
        for k in range(1, len(data)):      
            distancesKNN = allSortedDistances[:k]
            labelsKNN = np.array(distancesKNN)[:,-1]
            correctLabelsCounted = (labelsKNN == currentStyle).sum()      
            total = len(distancesKNN)
            accuracy = correctLabelsCounted / total
            currentK.append([accuracy, k, currentStyle])
            if accuracy > highestAccuracyAtK[k]:
                highestAccuracyAtK[k] = accuracy
        # once we reach the maximum number of labels in the data set, accuracy will only decline from here, so we break
            if correctLabelsCounted == maxTrainingLabelsCount:
                break

        currentK = sorted(currentK, reverse = True)
        if currentK[0][1] not in K and currentK[0][0] > accuracyLimit:
            K.append(int(currentK[0][1]))
    # here we want to get the most common accuracy for each k in K from frequencyOfAccuracyAtEachK.      
    fig, ax = plt.subplots()
    accLength = 100
    ax.plot(range(1, accLength), highestAccuracyAtK[1:accLength])
    ax.set(xlabel = 'k values', ylabel = 'accuracy at k', title = 'Accuracy of a value k during Training')
    ax.grid()
    fig.savefig('Accuracy of a value k during Training.png')
    plt.show()
    K = sorted(K)
    print('\nOptimized K values based on highest accuracy: ')
    print(K)
    return K
    

#***************************************************************************************
# function to test our optimized K values

def testOptimizedKUsingNumpy(data, style, K):
    ''' Use ideal values of k on test set of BeerClass objects

    Parameters
    ----------
    data : numpy array
        dataset of all BeerClass object\'s features matrix in numpy format
        
    style : numpy array
        dataset of all BeerClass object\' categorys keys in numpy format

    K : list
        ideal values of k derived from testing data set

    '''

    # results will be [[k, highest accuracy, average accuracy, lowest accuracy]]    
    highestAccuracyAtK = []
    for k in K:
        highestAccuracyAtK.append([k, float(0)])
#    indexCounter = -1
    for currentData, currentStyle in zip(data, style):
        allSortedDistances = getSortedDistancesUsingNumpy(data, currentData, style)[1:]
        currentK = []
        accuracyIndexer = -1
        for k in K:
            accuracyIndexer += 1              
            distancesKNN = allSortedDistances[:k]            
            labelsKNN = np.array(distancesKNN)
            correctLabelsCounted = (labelsKNN == currentStyle).sum()                  
            total = len(distancesKNN)
            accuracy = correctLabelsCounted / total    
            currentK.append([accuracy, k])
            if highestAccuracyAtK[accuracyIndexer][1] < accuracy:
                highestAccuracyAtK[accuracyIndexer][1] = float(accuracy)
        currentK = sorted(currentK, reverse = True)        
    print('\nHighest accuracies at K')
    print(np.array(highestAccuracyAtK))
    fig, ax = plt.subplots()
    accLength = 40
    x = np.array(highestAccuracyAtK)[:, 0]
    y = np.array(highestAccuracyAtK)[:, 1]
    ax.plot(x, y)
    ax.set(xlabel = 'k values', ylabel = 'accuracy at k', title = 'Testing Accuracy of optimized k values')
    ax.grid()
    fig.savefig('Testing Accuracy of optimized k values.png')
    plt.show()
    K = sorted(K)
    print('\nOptimized K values based on highest accuracy: ')
    print(K)

#**************************************************************************************************
# Function to classify a new beer based on optimized K nearest neigbors.

def classifyANewBeerUsingNumpy(aBeerFeatures, aBeerStyles, uBeerFeatures, K, name, classDictionary):
    '''Use dataset formatted as numpy arrays created in dataSetup to find optimized values of k for kNN algorithm

    Parameters
    ----------
    aBeerFeatures : numpy array
        dataset of all BeerClass object\'s features matrix in numpy format
        
    aBeerStyles : numpy array
        dataset of all BeerClass object\' categorys keys in numpy format

    uBeerFeatures : numpy array
        dataset of a new BeerClass object\'s features matrix in numpy format

    K : int
        ideal values of k derived from testing data set

    name : str
        name of new BeerClass object being classified

    classDictionary : dict
        key-values of category keys as keys and category names as values

    Returns
    ----------
    mostCommonStylesAtK : list
        list of the most frequent or common styles in k-nearest neighbors

    '''
    
    allSortedDistances = getSortedDistancesUsingNumpy(aBeerFeatures, uBeerFeatures, aBeerStyles)
    accuracyIndexer = -1
    mostCommonStylesAtK = []
    graphingData = {} 
    for k in K:
        distancesKNN = allSortedDistances[:k]
        labelsKNN = np.array(distancesKNN) 
        mostLabelsInDistancesKNN = Counter(labelsKNN[:,1]).most_common(1)
        mostCommonStylesAtK.append([k, mostLabelsInDistancesKNN[0][0], distancesKNN])
        dist = []
        style = []
        for i in distancesKNN:
            dist.append(i[0])
            style.append(i[1])
        graphingData.update({k: [dist, style]})
    __graphClassifications(graphingData, name, classDictionary)
    return mostCommonStylesAtK

#*****************************************************************************************************************
# function that graphs the features of a new beer and the features using colors to represent classifications of the k nearest neighbors. 

def __graphClassifications(data, title, classDictionary):
    
    color = ['red', 'green', 'blue', 'yellow', 'cyan', 'magenta', 'brown', 'coral', 'darkgreen', 'gold', 'fuchsia', 'lightblue', 'maroon', 'teal', 'violet']
    # here we unpack our data.
    x_kValues = []
    y_distances = []
    z_classifications = []
    z_classificationsColor = {}
    for k, dist_style in data.items():
        x_kValues.append(k)
        y_distances = [d for d in dist_style[0]]
        z_classifications = [c for c in dist_style[1]]

    
    uniqueClassifications = list(set(z_classifications))
    colorIndex = 0
    for i in uniqueClassifications:
        z_classificationsColor.update({i: [color[colorIndex], classDictionary[i]]})
        colorIndex += 1

    fig, ax = plt.subplots()
    for x in range(0, len(x_kValues)):
        classificationIndex = int(z_classifications[x])
        label = z_classificationsColor[classificationIndex][1]
        classColor = z_classificationsColor[classificationIndex][0]
        ax.scatter(x_kValues[x], y_distances[x], s = 20, label = label, color = classColor)
        plt.plot([x + 1, x_kValues[x]], [0, y_distances[x]], color = classColor)

    plt.title(title)
    legend = ax.legend(bbox_to_anchor=(1.05, 1), loc = 'upper left')
    ax.set(xlabel = 'k values', ylabel = 'distance')
    fig.savefig('Classification of ' + title + '.png')
    plt.show()
    
#*******************************************************************************************************************************
# Main Menu Options
def main():
    ''' Starts a menu format for running classification and recommendation '''
   
    # empty dictionaries of recommendations based on neighbors to userPredict based on different sets of features.
    # our lists of beer objects
    userInput = userCrossValidation = allBeer = []
    categoryDictionary = {}
    optimizedK = None
    knnResult = []
    knnResultingLabel = None    
    choice = -1
    while choice < 0 or choice > 2:
        print('\nData Options')
        print('0. Go back: this will erase all loaded data')
        print('1. Load Beer Information: load collected beer, user input, and cross validation information')
        print('2. KNN Operations: options for running KNN operations')


        try:
            choice = int(input())
        except:
            print('\nInvalid choice, please choose wisely.')
        if choice == 1:
            print('\nLoading Beer Information..')
            
            allBeer, userInput, userCrossValidation, categoryDictionary = loadBeerInformation(getUser = True, getCategoryDictionary = True)
            AllBeerDict = {}    # dictionary of all beer, with the beer's unique key being the key, and the beer's features array being the value
            for beer in allBeer:
                AllBeerDict[beer.getBeerKey()] = beer.getBeerFeaturesMatrix()
            userInputDict = {}  # dictionary of user input beer, with the beer's unique key being the key, and the beer's features array being the value
            for userP in userInput:
                userInputDict[userP.getBeerKey()] = userP.getBeerFeaturesMatrix()
            allNNResult = {}    
            
            print('\nLoading Complete')


        elif choice == 2:           
            
            kChoice = -1
            while kChoice < 0 or kChoice > 2:
                print('\nK Nearest Neighbor Options')
                print('0. Go back: this will erase all KNN-related data')
                print('1. Optimize K: function to optimize K')
                print('2. Classify user input using KNN algorithm and optimized K')
                print('3. Run All Nearest Neighbors algorithm')
                print('4. Make Recommendations based on B Nearest Neighbors')
                try:
                    kChoice = int(input())
                except:
                    print('\nInvalid choice, please choose wisely.')

### kChoice == 1: Optimize K
                if kChoice == 1:
                    
                    np.set_printoptions(threshold=np.inf, precision = 3)
                    # Train
                    # Test
                # so we are going to divide our data into training and test sets
                    trainBeerStyleKeys = [0] * 150
                    trainBeer = []
                    testBeerStyleKeys = [0] * 150
                    testBeer = []

                    maxTrainingLabelsCount = 40
                    maxTestingLabelsCount = 10
                    
                    for i in range(0, len(allBeer)):
                        if trainBeerStyleKeys[(allBeer[i].getBeerCategoryKey() - 1)] < maxTrainingLabelsCount:
                            trainBeer.append(allBeer[i])
                            trainBeerStyleKeys[allBeer[i].getBeerCategoryKey() - 1] += 1
                        elif testBeerStyleKeys[(allBeer[i].getBeerCategoryKey() - 1)] < maxTestingLabelsCount:
                            testBeer.append(allBeer[i])
                            testBeerStyleKeys[allBeer[i].getBeerCategoryKey() - 1] += 1
                            
                    shuffledTrainBeer = trainBeer            
                    random.shuffle(shuffledTrainBeer)
                    shuffledTestBeer = testBeer            
                    random.shuffle(shuffledTestBeer)
     
                    trainSetFeatures_X, trainSetStyles_y = dataSetup(trainBeer, yAvailable = True)
                    testSetFeatures_X, testSetStyle_y = dataSetup(testBeer, yAvailable = True)

                    print('\nOptimizing K')
                    optimizedK = optimizeKUsingNumpy(trainSetFeatures_X, trainSetStyles_y, maxTrainingLabelsCount)   # this is using list of type [numpy.array(), beerKey], to save time.
                    print('\nOptimizing Complete!')

                    print('\nTesting Optimized K')
                    testOptimizedKUsingNumpy(testSetFeatures_X, testSetStyle_y, optimizedK)
                    print('\nTesting Complete!')

### kChoice == 2: Classify User Input
                elif kChoice == 2:

                    
                    if len(optimizedK) != 0:
                        # this is where we convert allBeer into data and userPredict and userCrossValidation into data sets
                        # we will use allBeer data and optimizedK sets to classify each userPredict and userCrossValidation
                        # we must then graph the results, red for userPredict classification and black for the other classifications??

                        allBeerFeatures_X, allBeerStyles_y = dataSetup(allBeer, yAvailable = True)
                        for beer in userInput:
                            recommendationResultToGraph = {}
                            userBeerFeatures_X = dataSetup([beer])
                            listOfKResults = classifyANewBeerUsingNumpy(allBeerFeatures_X, allBeerStyles_y, userBeerFeatures_X, optimizedK, beer.getBeerName(), categoryDictionary)
                            print('\nBeer Name:      ' +  beer.getBeerName())
                            print('Beer Classifications based on optimized K values:')
                            for kResults in listOfKResults:
                                print('\nK-value:        ')
                                print(kResults[0])
                                print('Style Name:     ')
                                print(categoryDictionary[int(kResults[1])])
                    else:
                        print('\nIt might be helpful to run option 1 first..')

#  Option to run allNNResults for user input and cross validation.

                if kChoice == 3:                    
                    try:
                        # calculate all nearest neighbor distances
                        print('\nRunning all nearest neighbors algorithm on all features..')
                        allNNResult = calculateAllNN(AllBeerDict, userInputDict)
                        print('\nAll nearest neighbors  calculation complete!')
                    except:
                        print('You need to load your beer data first!')
#  Option to print the allNNResults
                if kChoice == 4:
                    print('\nHow many B recommendations do you want for each input? ')
                    try:
                        b = int(input())
                    except:
                        print('\nInvalid choice, please choose wisely.')
                    print('\nPrinting B Recommendations:')                    
                    printBNearestNeighbors(allNNResult, allBeer, b, userInput)
                    print('\nRecommendations complete!')
                    
                if kChoice != 0:
                    kChoice = -1
                else:
                    break
                
        if choice != 0:
            choice = -1  
        else:
            break

if __name__ == "__main__":
    main()
