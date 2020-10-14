from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import BeerCategoryClass
import BeerClass
import difflib
import fnmatch
import os
import pandas as pd
import re
import time

# funtion where selenium gathers html from each web page
def seleniumGetsHTML(site):

    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.get(site)
    time.sleep(3)  ### Timer to allow time for the compiler to grab html

    html = BeautifulSoup(driver.page_source, 'html.parser')

    driver.close()
    driver.quit()

    return html

## returns the text of the starting node.
# add the first node to the tree here...
def buildTree(html, site):
    
    # create a root node
    rootNodeOfTree = BeerCategoryClass.BeerCategory()   
    key = 0
    # give root node basic attributes
    rootNodeOfTree.setCategoryName('All Beers')
    rootNodeOfTree.setCategoryKey(key)              # root node will have a key of 0 every time.
    rootNodeOfTree.setCategoryParent(None)          # root node has no parents
    rootNodeOfTree.setCategoryParentKey(key - 1)    # root node does not have a parent, thus root node's parent key is set to -1
    rootNodeOfTree.setCategory_href('/beer/styles') # since we are starting that this page manually, we will go ahead and enter this in manually
    rootNodeOfTree.setSubCategoriesExist(True)
    
    styleBreakClasses = html.findAll('div', {'class': 'stylebreak'})  # this will get all of the div tags with class 'stylebreak'
    for styleBreak in styleBreakClasses:
        key += 1
        currentCategory = BeerCategoryClass.BeerCategory()
        currentCategory.setCategoryName(styleBreak.find('b').get_text())
        currentCategory.setCategoryKey(key)
        currentCategory.setCategoryParent(rootNodeOfTree)
        currentCategory.setCategoryParentKey(rootNodeOfTree.getCategoryKey())
        currentCategory.setSubCategoriesExist(True)
        subCategoryLists = styleBreak.findAll('a')
        for subCategory in subCategoryLists:
            key += 1
            currentSubCategory = BeerCategoryClass.BeerCategory()
            currentSubCategory.setCategoryName(subCategory.get_text())
            currentSubCategory.setCategoryKey(key)
            currentSubCategory.setCategoryParent(currentCategory)
            currentSubCategory.setCategoryParentKey(currentCategory.getCategoryKey())
            currentSubCategory.setCategory_href(subCategory['href'])
            currentCategory.addSubCategory(currentSubCategory)
        rootNodeOfTree.addSubCategory(currentCategory)
    return rootNodeOfTree

    
#******************************************************************************************************************************
# Outputting tree to screen
# prints a visual representation of the categories and sub categories tree.
def printCategoryTree(currentCategory, level):

    if currentCategory.doSubCategoriesExist() == True:
        for sub in currentCategory.getSubCategories():
            print((' |    ')*level)
            print((' |    ')*level)
            print((' |    ')*(level-1) + ' |--' + sub.getCategoryName() + '  Key: ' + str(sub.getCategoryKey()))
            printCategoryTree(sub, (level+1))
        level -= 1    

#*******************************************************************************************************
# Sending tree information to an excel sheet.


def createWorkbook(root):

    for category in root.getSubCategories():
        for subCategory in category.getSubCategories():
            wb = Workbook()
            categoryName = re.sub('\/', 'and', subCategory.getCategoryName())
            addToNewWorkbook(wb, subCategory, 'None', -1, index = 0)
            fileName = 'Beer\\'
            openFile = categoryName + '.xlsx'
            wb.save(openFile)

            # change this so that each beer has its own page.

def addToNewWorkbook(wb, currentCategory, pName, pkey, index):
    currentRow = 0
    currentColumn = 0

    categoryName = re.sub('\/', 'and', currentCategory.getCategoryName())
    wb.create_sheet(index = index, title = categoryName)
    wb.active = index
    sheet = wb.active
    
    catName = sheet.cell(row = 1, column = 1)
    catName.value = 'Category Name:'
    catName = sheet.cell(row = 1, column = 2)    
    catName.value = currentCategory.getCategoryName()

    catKey = sheet.cell(row = 2, column = 1)
    catKey.value = 'Category Key:'
    catKey = sheet.cell(row = 2, column = 2)    
    catKey.value = currentCategory.getCategoryKey()

    catParent = sheet.cell(row = 3, column = 1)
    catParent.value = 'Category Parent:'
    catParent = sheet.cell(row = 3, column = 2)    
    catParent.value = currentCategory.getCategoryParent().getCategoryName()
    
    catParentKey = sheet.cell(row = 4, column = 1)
    catParentKey.value = 'Category Parent Key:'
    catParentKey = sheet.cell(row = 4, column = 2)    
    catParentKey.value = currentCategory.getCategoryParentKey()
        
    catHREF = sheet.cell(row = 5, column = 1)
    catHREF.value = 'Category href:'
    catHREF = sheet.cell(row = 5, column = 2)
    catHREF.value = currentCategory.getCategory_href()

    catDesc = sheet.cell(row = 6, column = 1)
    catDesc.value = 'Category Description:'
    catDesc = sheet.cell(row = 6, column = 2)    
    catDesc.value = currentCategory.getCategoryDescription()

    catDesc = sheet.cell(row = 7, column = 1)
    catDesc.value = 'Minimum ABV:'
    catDesc = sheet.cell(row = 7, column = 2)    
    catDesc.value = currentCategory.getCategoryMinABV()

    catDesc = sheet.cell(row = 8, column = 1)
    catDesc.value = 'Maximum ABV:'
    catDesc = sheet.cell(row = 8, column = 2)    
    catDesc.value = currentCategory.getCategoryMaxABV()

    catDesc = sheet.cell(row = 9, column = 1)
    catDesc.value = 'Minimum IBU:'
    catDesc = sheet.cell(row = 9, column = 2)    
    catDesc.value = currentCategory.getCategoryMinIBU()

    catDesc = sheet.cell(row = 10, column = 1)
    catDesc.value = 'Maximum IBU:'
    catDesc = sheet.cell(row = 10, column = 2)    
    catDesc.value = currentCategory.getCategoryMaxIBU()

    ## category features go here.
    currentRow = 15
    if len(currentCategory.getCategoryFeaturesMatrix()) == 18:
        i = 0
        features = eachItem.getBeerFeaturesMatrix()
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'ABV'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Astringency'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Body'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Alcohol'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Bitter'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Sweet'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Sour'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Salty'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Dark Fruits'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Citrus Fruits'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Tropical Fruits'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Hoppy'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Floral'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Spicy'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Herbal'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Malty'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Richness'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Yeast'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1

    # save a list of beer and their review count,
    # will use this to determine how impacting the reviews will be on features
    var = sheet.cell(row = 40, column  = 1)
    var.value = 'Beer Name'
    var = sheet.cell(row = 40, column  = 2)
    var.value = 'Review Count'    
    currentRow = 40
    currentColumn = 1
    for beer in currentCategory.getCategoryBeers():
        var = sheet.cell(row = currentRow, column  = 1)
        var.value = beer.getBeerName()
        var = sheet.cell(row = currentRow, column  = 2)
        var.value = len(beer.getBeerReviewsFullContent())
        currentRow += 1

    # enter information for each beer into their own sheets after category information
    for eachItem in currentCategory.getCategoryBeers():
        index += 1
        beerName = eachItem.getBeerName()
        beerName = re.sub('[^A-Za-z0-9]+', '', beerName)
        wb.create_sheet(index = index, title = beerName)
        wb.active = index
        sheet = wb.active
        
        label = sheet.cell(row = 1, column = 1)
        label.value = 'Beer Name'
        var = sheet.cell(row = 1, column = 2)
        var.value = eachItem.getBeerName()
        
        label = sheet.cell(row = 2, column = 1)
        label.value = 'Beer key'
        var = sheet.cell(row = 2, column = 2)
        var.value = eachItem.getBeerKey()
        
        label = sheet.cell(row = 3, column = 1)
        label.value = 'Beer Style'
        var = sheet.cell(row = 3, column = 2)
        var.value = eachItem.getBeerStyle()
        
        label = sheet.cell(row = 4, column = 1)
        label.value = 'Beer Style Key'
        var = sheet.cell(row = 4, column = 2)
        var.value = eachItem.getBeerCategoryKey()
        
        label = sheet.cell(row = 5, column = 1)
        label.value = 'Brewery'
        var = sheet.cell(row = 5, column = 2)
        var.value = eachItem.getBeerBrewery()
        
        label = sheet.cell(row = 6, column = 1)
        label.value = 'Ave Rating'
        var = sheet.cell(row = 6, column = 2)
        var.value = eachItem.getBeerAverageRating()
        
        label = sheet.cell(row = 7, column = 1)
        label.value = 'Beer ABV'
        var = sheet.cell(row = 7, column = 2)
        var.value = eachItem.getBeerABV()
        
        label = sheet.cell(row = 8, column = 1)
        label.value = 'Beer Min IBU'
        var = sheet.cell(row = 8, column = 2)
        var.value = eachItem.getBeerMinIBU()
               
        label = sheet.cell(row = 9, column = 1)
        label.value = 'Beer Max IBU'
        var = sheet.cell(row = 9, column = 2)
        var.value = eachItem.getBeerMaxIBU()
        
        label = sheet.cell(row = 10, column = 1)
        label.value = 'Beer Description'
        var = sheet.cell(row = 10, column = 2)
        try:
            var.value = eachItem.getBeerDescription()
        except:
            var.value = 'error entering this description'
        
        label = sheet.cell(row = 20, column = 1)
        label.value = 'Reviews:'
        currentRow = 21
        currentColumn = 1
        columnWidth = 100
        for eachReview in eachItem.getBeerReviewsFullContent():
            var = sheet.cell(row = currentRow, column = currentColumn)            
            try:
                var.value = eachReview
            except:
                var.value = 'error entering this review'
            currentRow += 1

        # this is where we will save our beer features values
        currentRow += 5
        if len(eachItem.getBeerFeaturesMatrix()) == 18:
            i = 0
            features = eachItem.getBeerFeaturesMatrix()
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'ABV'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Astringency'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Body'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Alcohol'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Bitter'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Sweet'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Sour'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Salty'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Dark Fruits'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Citrus Fruits'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Tropical Fruits'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Hoppy'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Floral'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Spicy'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Herbal'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Malty'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Richness'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Yeast'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]


        # word count for each beer from description and reviews saved here
        currentRow += 5
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Words'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = 'Counts'
        
        currentRow += 1
        currentColumn = 1
        for beer in currentCategory.getCategoryBeers():
            if beer.getBeerWordCount() != None:
                for keyVal in beer.getBeerWordCount():
                    currentCell = sheet.cell(row = currentRow, column = currentColumn)
                    # sends word to excel file
                    variable = currentCell
                    variable.value = keyVal[0]
                    currentColumn += 1
                    currentCell = sheet.cell(row = currentRow, column = currentColumn)
                    # sends word count to excel file#
                    variable = currentCell
                    variable.value = keyVal[1]
                    currentRow += 1
                    currentColumn = 1
        

# Look for any empty or blank pages and remove them from the workbook (this usually occurs at the end of the workbook)
    for sheet in wb:
        if sheet.cell(row = 1, column = 1).value == '' or sheet.cell(row = 1, column = 1).value == None:
            wb.remove(sheet)

#*******************************************************************************************************
# Get information from excel file, given that there is an implemented tree already made and information
# just needs to be filled in

def loadSubCategories(tree):
           
    fileDirectory = 'D:\Python Projects\Beer Recommender Project\Beer'
    # first we need to see if the name of the categories we want to collect data from exist as files
    fileList = os.listdir(fileDirectory)
    excelFileList = []
    fileName = 'Beer\\'
    for file in fileList:
        excelFileList.append(fileName + file)  # list of file names

    for excel in excelFileList:
        keyFound = False
        wb = load_workbook(excel)
        wb.active = 0
        key = int(wb.active.cell(row = 2, column = 2).value)
        for category in tree.getSubCategories():
            for subCategory in category.getSubCategories():
                if subCategory.getCategoryKey() == key:
                    print(subCategory.getCategoryName())
                    subCategory = gatherInformation(0, wb, subCategory)
                    subCategory.setCategoryParent(category)
                    keyFound == True
                    break
            if keyFound == True:
                break

    category = tree.getSubCategories()[0]
#    for subCategory in category.getSubCategories():
#        print(subCategory.getCategoryName())
#        print(subCategory.getCategoryMaxIBU())
                    
    print('loading complete')

    return tree

def gatherInformation(index, wb, tempCategory):
    wb.active = index
    sheet = wb.active
    tempCategory.setCategoryName(sheet.cell(row = 1, column = 2).value)
    tempCategory.setCategoryKey(int(sheet.cell(row = 2, column = 2).value))
    # we set category parent to the object being iterated after this function
    tempCategory.setCategoryParentKey(int(sheet.cell(row = 4, column = 2).value))
    tempCategory.setCategory_href(sheet.cell(row = 5, column = 2).value)
    tempCategory.setCategoryDescription(sheet.cell(row = 6, column = 2).value)
    tempCategory.setCategoryMinABV(float(sheet.cell(row = 7, column = 2).value))
    tempCategory.setCategoryMaxABV(float(sheet.cell(row = 8, column = 2).value))
    tempCategory.setCategoryMinIBU(int(sheet.cell(row = 9, column = 2).value))
    tempCategory.setCategoryMaxIBU(int(sheet.cell(row = 10, column = 2).value))
    # this is where we will load our keyword data bank for this beer
    currentRow = 15
    i = 0
    features = []
#    while i < 18:
#        var = sheet.cell(row = currentRow, column = 2)
#        if var.value != '' or var.value != None:
#            features.append(sheet.cell(row = currentRow, column = 2).value)
#        currentRow += 1
#        i += 1
#    tempCategory.setCategoryFeaturesMatrix(features)    
    index += 1
    
    while index < len(wb.worksheets):
        wb.active = index
        sheet = wb.active
        
        currentColumn = 2
        currentRow = 1
        currentCell = sheet.cell(row = currentRow, column = currentColumn)
        while currentCell.value != '' and currentCell.value != None:
            tempItem = BeerClass.Beer()

            value = currentCell.value
            tempItem.setBeerName(value)
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            
            value = currentCell.value
            if value == '' or value == None:
                value = 0
            tempItem.setBeerKey(int(value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            
            value = currentCell.value
            if value == '' or value == None:
                value = 'NA'
            tempItem.setBeerStyle(value)
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            
            value = currentCell.value
            if value == '' or value == None:
                value = 0
            tempItem.setBeerCategoryKey(int(value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            
            value = currentCell.value
            if value == '' or value == None:
                value = 'NA'
            tempItem.setBeerBrewery(value)
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            
            value = currentCell.value
            if value == '' or value == None:
                value = 0.0
            tempItem.setBeerABV(float(value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            
            value = currentCell.value
            if value == '' or value == None:
                value = 0.0
            tempItem.setBeerAverageRating(float(value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            
            value = currentCell.value
            if value == '' or value == None:
                value = 0
            tempItem.setBeerMinIBU(int(value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            
            value = currentCell.value
            if value == '' or value == None:
                value = 0
            tempItem.setBeerMaxIBU(int(value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            
            value = currentCell.value
            if value == '' or value == None:
                value = 'NA'
            tempItem.setBeerDescription(value)

            # gets each review from the file
            currentRow = 21
            currentColumn = 1
            while currentCell.value != '' and currentCell.value != None:
                value = currentCell.value
                tempItem.addBeerReviewsFullContent(value)
                currentRow += 1
                currentCell = sheet.cell(row = currentRow, column = currentColumn)
            tempCategory.addCategoryBeer(tempItem)

            # this is where we will load our keyword data bank for this beer
            currentRow += 5
            i = 0
            features = []
#            while i < 18:
#                var = sheet.cell(row = currentRow, column = 2)
#                print(var.value)
#                if var.value != '' or var.value != None:
#                    features.append(sheet.cell(row = currentRow, column = 2).value)
#                currentRow += 1
#                i += 1
#            tempItem.setBeerFeaturesMatrix(features)
            
            # here we need to get word counts
            
            
        index += 1

        

    return tempCategory

#*************************************************************************************************************
## these sets of functions web scrapes heb.com in each category and loads items to their respective categories
## in the category tree.

def startGetCategoryItems(root, site):

    sinceEpoch = time.time()
    startTimeObj = time.localtime(sinceEpoch)
    driver = webdriver.Chrome(ChromeDriverManager().install())
    currentSite = site + root.getCategory_href()
    driver.get(currentSite)
    time.sleep(3)

    # we know that there are no items in each category in root's list of sub categories
    # we move on to each category's sub category where we will find our beer.

    # first we will see if there are any category files already in existence due to interrupts    
    key = 0
    for eachCategory in root.getSubCategories():
        for eachSubCategory in eachCategory.getSubCategories():
            # this is done to skip any category scraping for data we already have after loading
            # if we have more than one beer in this eachSubCategory, then we know that we already
            # have this information from a previous scrape session.
            # this has to be implemented because of SPECTRUM's wonderful internet connection....
            # to save time
            if len(eachSubCategory.getCategoryBeers()) == 0:
                print('Scraping ' + eachSubCategory.getCategoryName())
                key = openSubCategoryPages(eachSubCategory, driver, site, key)
    
    driver.close()
    driver.quit()
    
    sinceEpoch = time.time()
    endTimeObj = time.localtime(sinceEpoch)
    
    print('Started: %d:%d' %(startTimeObj.tm_hour, startTimeObj.tm_min))
    print('Finished: %d:%d' %(endTimeObj.tm_hour, endTimeObj.tm_min))

    return root

def cleanupDoubleDecimalStrings(string):

    if string != '' or string != None:
        newString = re.sub('[a-zA-Z()\s$\/%:]', '', string)
        minIBU = maxIBU = ''
        dividerFound = False
        for c in newString:
            if c == '-':
                dividerFound = True
                continue
            if dividerFound == False:
                minIBU += c
            else:
                if c == '|' or c == '%':
                    break
                maxIBU += c
    return minIBU, maxIBU

def cleanupSingleDecimalStrings(string):

    if string != '' or string != None:
        newString = re.sub('[a-zA-Z()\s$/%:]', '', string)
    return newString
    

def openSubCategoryPages(currentCategory, dr, site, key):

    # open pages of sub categories to get to list of beer   
    currentSite = site + currentCategory.getCategory_href()
    dr.execute_script("window.open(''); ")
    dr.switch_to.window(dr.window_handles[1])
    dr.get(currentSite)
    time.sleep(3)

    allHTML = BeautifulSoup(dr.page_source, 'html.parser')

    # details of beer style can be found in first <div> tag of <div id="ba-content">
    detailsTag = allHTML.find('div', {'id' : 'ba-content'})
    detailsTag = detailsTag.findAll('div')[0]
    currentCategory.setCategoryDescription(detailsTag.get_text())
    decimals = detailsTag.findAll('span')
    minABV, maxABV = cleanupDoubleDecimalStrings(decimals[0].get_text())
    minIBU, maxIBU = cleanupDoubleDecimalStrings(decimals[1].get_text())
    try:
        currentCategory.setCategoryMinABV(float(minABV))
        currentCategory.setCategoryMaxABV(float(maxABV))
    except:
        print('had an issue getting ABV in ' + currentCategory.getCategoryName())
        currentCategory.setCategoryManualEditFlag(True)
    try:
        currentCategory.setCategoryMinIBU(float(minIBU))
        currentCategory.setCategoryMaxIBU(float(maxIBU))
        currentCategory.setCategoryManualEditFlag(True)
    except:
        print('had an issue getting IBU in ' + currentCategory.getCategoryName())
    
    # only one tbody tag that holds line items of beer
    tbodyTag = allHTML.find('tbody')
    trTags = tbodyTag.findAll('tr')
    # the first three <tr> tags are as follows
    # tag 1: non-important
    # tag 2: links to the next 50 beers, not important yet.
    # tag 3: table headers allowing for re-ordering of beers, not very important
    # tag 4: start of beers, very important.
    # ...
    # ...
    # tag last: links to get to next 50 beers, very important.
    # starting at index 3, we will get the beer information available on this page.

    # this is where we go to each beer's individual website.
    
    index = 3 
    maxCount = len(trTags) - 2
    while index < maxCount:
        key += 1
        current_href = trTags[index].find('a')['href']
        # open pages of sub categories to get to list of beer   
        currentSite = site + current_href
        dr.execute_script("window.open(''); ")
        dr.switch_to.window(dr.window_handles[2])
        dr.get(currentSite)
        time.sleep(3)
        thisBeerHTML = BeautifulSoup(dr.page_source, 'html.parser')
        thisBeer = BeerClass.Beer()
        # since beerName is name of beer and name of brewery on the same string,
        # we get the the breweryName and sub out of beerName
        beerName = thisBeerHTML.find('div', {'class': 'titleBar'})
        breweryName = beerName.find('span').get_text()
        beerName = beerName.get_text()
        beerName = re.sub(breweryName, '', beerName)
        thisBeer.setBeerName(beerName)
        thisBeer.setBeerBrewery(breweryName)
        thisBeer.setBeerKey(key)
        thisBeer.setBeerCategoryKey(currentCategory.getCategoryKey())
        thisBeer.setBeerStyle(currentCategory.getCategoryName())
        # for finding beer stats
        beerStats = thisBeerHTML.findAll('dd', {'class': 'beerstats'})
        abv = cleanupSingleDecimalStrings(beerStats[1].get_text())
        try:
            thisBeer.setBeerABV(float(abv))
        except:
            print(thisBeer.getBeerName() + ' does not have an ABV on the website.\nSubstituting with the category average ABV')
            try:
                thisBeer.setBeerABV((float(minABV) + float(maxABV)) / 2)
            except:
                thisBeer.setBeerManualEditFlag(True)
        thisBeer.setBeerAverageRating(float(beerStats[3].find('span', {'class': 'ba-ravg Tooltip'}).get_text()))
            
        # for finding notes
        notes = thisBeerHTML.find('div', {'style': 'clear:both; margin:0; padding:0px 20px; font-size:1.05em;'})
        thisBeer.setBeerDescription(notes.get_text())
        try:
            thisBeer.setBeerMinIBU(float(minIBU))
            thisBeer.setBeerMaxIBU(float(maxIBU))
        except:
            thisBeer.setBeerManualEditFlag(True)
        # for finding all beer reviews
        beerReviews = thisBeerHTML.findAll('div', {'id': 'rating_fullview_content_2'})
        for each in beerReviews:
            thisBeer.addBeerReviewsFullContent(each.get_text())
        print(thisBeer.getBeerName() + "   " + str(len(thisBeer.getBeerReviewsFullContent())))
        
        currentCategory.addCategoryBeer(thisBeer)
    
        dr.close()
        dr.switch_to.window(dr.window_handles[1])
        index += 1

        # this is where we get the next button to move to the second page.
#        pageCount += 1
#        nextPage = site + (trTags[-1].find('a')['href'])
#        dr.get(nextPage)
#        time.sleep(3)
#        allHTML = BeautifulSoup(dr.page_source, 'html.parser')
    
    dr.close()
    dr.switch_to.window(dr.window_handles[0])

    # save to a new excel document after each category is scraped
    # because SPECTRUM...
    wb = Workbook()
    addToNewWorkbook(wb, currentCategory, currentCategory.getCategoryName(), currentCategory.getCategoryKey(), index = 0)
    name = re.sub('\/', 'and',currentCategory.getCategoryName())
    file = 'Scrape\\'
    openFile = file + name + '.xlsx'
    wb.save(openFile)


    return key

#*******************************************************************************************************************************
# Manual Key Reassignment to individual beers so that each beer is garaunteed a unique key

def reassignKeys(root):
    key = 0
    for category in root.getSubCategories():
        for subCategory in category.getSubCategories():
            for beer in subCategory.getCategoryBeers():
                key += 1
                beer.setBeerKey(key)
    return root

#**********************************************************************************************************************************
# set of functions build keyword data banks

def compileWordCounts(root):

    invalidWords = []
    file = open('omitted words.txt', 'r')
    for word in file:
        word = re.sub('\n', '', word)
        invalidWords.append(word)

    # dictionary for all of the words in the category.
    # it gets broken down into sub categories
#    categoryWordCountDictionary = {}     # will house dictionaries in category
    subCategoryWordCountDictionary = {}  # will house dictionaries in sub category
    currentBeerWordCountDictionary = {}
    
    # for getting category word dictionaries of words
    for category in root.getSubCategories():
        
        # for getting sub category dictionaries of words
        for subCategory in category.getSubCategories():
            subCategoryWordCountDictionary = {}  # will house dictionaries in sub category
            # gets words from the description of the sub category
            # as stated by BeerAdvocate                    
            # for getting beer dictionaries of words
            for beer in subCategory.getCategoryBeers():
                currentBeerWordCountDictionary = {} # empty dictionary of word count for current beer
                # gets words from the description of the beer
                # as stated by the brewery
                allWords = beer.getBeerDescription()
                allWords = re.sub('[^A-Za-z]+', ' ', allWords)
                allWords = allWords.lower()
                currentWord = ''
                for c in allWords:
                    if c == ' ' and currentWord != '':
                        
                        isValid = True
                        for invalid in invalidWords:
                            if currentWord == invalid:
                                isValid = False
                                break
                            
                        if isValid == True:
                            if currentWord in currentBeerWordCountDictionary:
                                count = currentBeerWordCountDictionary[currentWord]
                                count += 1
                                currentBeerWordCountDictionary[currentWord] = count
                            else:
                                currentBeerWordCountDictionary[currentWord] = 1
                        currentWord = ''
                        
                    else:
                        currentWord += c
                        
                # then we get the words from all of the reviews
                # left by users on BeerAdvocate
                for review in beer.getBeerReviewsFullContent():
                    allWords = review
                    allWords = re.sub('[^A-Za-z]+', ' ', allWords)
                    allWords = allWords.lower()
                    currentWord = ''
                    for c in allWords:
                        if c == ' ' and currentWord != '':
                            
                            isValid = True
                            for invalid in invalidWords:
                                if currentWord == invalid:
                                    isValid = False
                                    break

                            if isValid == True:
                                break
                                if currentWord in currentBeerWordCountDictionary:
                                    count = currentBeerWordCountDictionary[currentWord]
                                    count += 1
                                    currentBeerWordCountDictionary[currentWord] = count
                                else:
                                    currentBeerWordCountDictionary[currentWord] = 1
                            currentWord = ''
                        else:
                            currentWord += c
                currentBeerWordCountDictionary = sorted(currentBeerWordCountDictionary.items(), key=lambda x: x[1], reverse=True)
                beer.setBeerWordCount(currentBeerWordCountDictionary)

            # here we will add all of the beer words to the sub category words so that we can build a profile of the categories as well
            for word in currentBeerWordCountDictionary:
                if word[0] in subCategoryWordCountDictionary:
                    count = subCategoryWordCountDictionary[word[0]]
                    count += word[1]
                    subCategoryWordCountDictionary[word[0]] = count
                else:
                    subCategoryWordCountDictionary[word[0]] = word[1]
            subCategory.setCategoryFeaturesMatrix(subCategoryWordCountDictionary)
    return root

#************************************************************************************************************************************
# function to save word counts to spreadsheets for each sub category

def sendWordCountToWorkBook(currentBeer, key, sheet):
    fileDirectory = 'D:\Python Projects\Beer Recommender Project\Beer'
    # first we need to see if the name of the categories we want to collect data from exist as files
    fileList = os.listdir(fileDirectory)
    excelFileList = []
    fileName = 'Beer\\'
    for file in fileList:
        excelFileList.append(fileName + file)  # list of file names
    
    # here we will rename each tab after the name of the beer
    # and insert the data into the workbook in some fasion..
    # Col A: word
    # Col B: word count
    
    variable = sheet.cell(row = 1, column = 1)
    variable.value = 'Beer Key'
    variable = sheet.cell(row = 1, column = 2)
    variable.value = key

    variable = sheet.cell(row = 2, column = 1)
    variable.value = 'Words'
    variable = sheet.cell(row = 2, column = 2)
    variable.value = 'Counts'
    
    currentRow = 3
    currentColumn = 1
    for keyVal in currentBeer:
        currentCell = sheet.cell(row = currentRow, column = currentColumn)
        # sends word to excel file
        variable = currentCell
        variable.value = keyVal[0]
        currentColumn += 1
        currentCell = sheet.cell(row = currentRow, column = currentColumn)
        # sends word count to excel file#
        variable = currentCell
        variable.value = keyVal[1]
        currentRow += 1
        currentColumn = 1
        
def saveWordCountsInCategory(root):

            wb = Workbook()
            for category in root.getSubCategories():
                for subCategory in category.getSubCategories():
                    index = 1
                    for beer in subCategory.getCategoryBeers():
                        beerName = beer.getBeerName()
                        beerName = re.sub('[^A-Za-z]+', ' ', beerName)
                        wb.create_sheet(index = index, title = beerName)
                        wb.active = index
                        sheet = wb.active                  
                        sendWordCountToWorkBook(beer.getBeerWordCount(), beer.getBeerKey(), sheet)
                        index += 1
                    fileName = 'Word Counts\\'
                    categoryName = re.sub('\/', 'and', subCategory.getCategoryName())
                    openFile = fileName + categoryName + '.xlsx'
                    wb.save(openFile)


#*************************************************************************************************************************************
# function to get a combined word count from all beers.  This is to assist with sifting for beer related words manually

def combineWordCounts(root):
    combineWordCount = {}

    for category in root.getSubCategories():
        for subCategory in category.getSubCategories():
            for beer in subCategory.getCategoryBeers():
                for wordCount in beer.getBeerWordCount():
                    if wordCount[0] in combineWordCount:
                        count = combineWordCount[wordCount[0]]
                        count += wordCount[1]
                        combineWordCount[wordCount[0]] = count
                    else:
                        combineWordCount[wordCount[0]] = wordCount[1]
    combineWordCount = sorted(combineWordCount.items(), key=lambda x: x[1], reverse=True)
                        
                            
    
    wb = Workbook()
    categoryName = 'Combine Word Count'
    fileName = 'Keyword Bank\\'
    openFile = fileName + categoryName + '.xlsx'
    wb.active = 0
    sheet = wb.active
    cell = sheet.cell(row = 1, column = 1)
    cell.value = 'Words'
    cell = sheet.cell(row = 1, column = 2)
    cell.value = 'Counts'
    
    currentRow = 2
    for wordCount in combineWordCount:
        cell = sheet.cell(row = currentRow, column = 1)
        cell.value = wordCount[0]
        cell = sheet.cell(row = currentRow, column = 2)
        cell.value = wordCount[1]
        currentRow += 1
    
    wb.save(openFile)

#***********************************************************************************************************************************
# function to build a features matrix for each beer by comparing word counts with those in the keyword data bank
# each count on a word will impact the order of magnitude for the feature it is stored under
# we will use difflib's get_close_matches(word, possibilities[, n][, cutoff]) with an n value of 3 or 4 close matches

def compileBeerFeaturesMatrix(root):
    n = 1
    cutoff = 0.7
    currentBeerWords = []
    missedWordsFromBeers = {}
    astringent = []
    body = []
    alcohol = []
    bitter = []
    sweet = []
    sour = []
    salty = []
    darkFruit = []
    citrusFruit = []
    tropicalFruit= []
    hoppy = []
    floral = []
    spicy = []
    herabl = []
    malty = []
    richness = []
    yeast = []
    fileName = 'Keyword Bank\\'
    file = 'Beer Descriptors.xlsx'
    wb = load_workbook(fileName + file)
    index = 0
    wb.active = index
    sheet = wb.active
    currentRow = 2
    currentColumn = 1
    currentCell = sheet.cell(row = currentRow, column = currentColumn)
    

    return tree


#***********************************************************************************************************************************
# menu options.


def treeMenuOptions():
    
    choice = -1
    while choice < 0 or choice > 8:
        print('\nBeer Recommender Project')
        print('What operation are we running?')
        print('0. QUIT')
        print('1. Build Tree: scrape website for its to build a tree of beer styles')
        print('2. Load Tree: AFTER building tree, load information from files\n   including all beer information into the tree')
        print('3. Print Category Tree: print the tree of beer styles and unique keys to screen')
        print('4. Save Tree: save the tree to multiple excel files')
        print('5. Scrape for Items: scrape website for items in each category\n   and add to the tree.  This also saves the data during the process')
        print('6. Manual Key Reassignment: reassigns keys to each beer\n   due to internet interruptions, multiple beers may have the same key')
        print('7. Word Count Options: display a list of options for word counts')
        print('8. Machine Learning Techniques: display a list of options for\n   machine learning techniques')
### Need a word counter put together with a means to save the word counter
### Need a key re-assignment for all beer, because we had to scrape data in multiple sessions there are multiple keys being reused.



# need a try except here.
        try:
            choice = int(input())
        except:
            print('\nInvalid choice, please choose wisely.')

        
    return choice

# option 1
# function that builds a category tree
def buildCategoryTree(url, name):
    print('\nBuilding category tree..\n')
    # our tree object where we append the nodes and their respective data values.
    tree = BeerCategoryClass.BeerCategory()
    # gets the html via selenium and beautifulsoup combination
    # since we are starting a '/beer/styles', we will go ahead and append that
    # here to get the html for that page.
    customStartingPoint = '/beer/styles'
    soupSource = seleniumGetsHTML(url + customStartingPoint)
    # get the tree's starting node.
    tree = buildTree(soupSource, url)
    print('\nTree has been built!')
    return tree

# option 2
# function to load existing beer information into a tree from workbooks
def loadFromWorkbook(tree):

    print('\nLoading beer information into initialized tree from workbooks..')
    tree = loadSubCategories(tree)
    return tree

# option 3
# function that prints a visual representation of the tree to the screen
def printTree(currentCategory):
    print('\nPrinting category tree to screen..\n')
    print(currentCategory.getCategoryName())
    printCategoryTree(currentCategory, 1)
    print('\nTree has been printed!')

# option 4
# function to save tree to a workbook
def saveToWorkbook(tree):

    print('\nSaving to file..')
    createWorkbook(tree)
    print('\nSaving complete!')

# option 5
# function to add items to each category in the tree
# usually the leaf nodes of a tree
def scrapeCategoryItems(tree, url):

    print('\nGathering item information from website ' + url + '..')
    tree = startGetCategoryItems(tree, url)
    print('\nGathering of item information Complete!')
    return tree

# option 6
# function that reassigns key values to each beer
# due to poor internet connection, multiple beer may have the same key.
def manualKeyReassignment(tree):
    print('\nStarting key reassignments to beer')
    tree = reassignKeys(tree)
    print('\nKey reassignment complete!')
    return tree

# option 7
# function to run a keyword bank builder based on reviews and descriptions
def wordBankBuilder(tree):

    wordBankChoice = -1
    while wordBankChoice != 0:
        wordBankChoice = -1
        print('0. Go Back')
        print('1. Compile Word Counts')
        print('2. Save Word Counts to a separate file')
        print('3. Save to combined word count file')
        print('4. Compile features matrix based on keyword data bank')
        try:
            wordBankChoice = int(input())
        except:
            print('\nInvalid choice, please choose wisely.')
        if wordBankChoice == 1:
            tree = compileWordCounts(tree)
        elif wordBankChoice == 2:
            saveWordCountsInCategory(tree)
        elif wordBankChoice == 3:
            combineWordCounts(tree)
        elif wordBankChoice == 4:
            compileBeerFeaturesMatrix(tree)
    return tree

# option 8
# displays the options available for machine learning techniques
def displayMachineLearningOptions(tree):
    choiceMachineLearning = -1
    while choiceMachineLearning != 0:
        choiceMachineLearning = -1
        print('0. Go Back')
        print('1. Compile Word Counts')
        print('2. Save Word Counts to a separate file')
        print('3. Save to combined word count file')
        print('5. ')
        print('6. ')
        print('7. ')
        try:
            choiceMachineLearning = int(input())
        except:
            print('\nInvalid choice, please choose wisely.')
#        if choiceMachineLearning == 1:
            
#        elif choiceMachineLearning == 2:
            
#        elif choiceMachineLearning == 3:
            


def testOption(root):

    for category in root.getSubCategories():
        for subCategory in category.getSubCategories():
            print(subCategory.getCategoryName())
            for beer in subCategory.getCategoryBeers():
                print('\n' + beer.getBeerName())
                print(str(len(beer.getBeerReviewsFullContent())))
                print(str(len(beer.getBeerWordCount)))
                print(beer.getBeerWordCount())
                        
#*******************************************************************************************************
# main function    
def main():

    # our tree object where we append the nodes and their respective data values.
    BeerCategoryTree = BeerCategoryClass.BeerCategory()

    # THIS ONLY WORKS FOR HEB.COM At the moment
    # page we will be visiting to build our tree
    beerURL = 'https://www.beeradvocate.com'
    # Grocery Name
    beerName = 'Beers'
    operationChoice = -1

    while(operationChoice != 0):
        if(operationChoice == 1):
            BeerCategoryTree = buildCategoryTree(beerURL, beerName)
            beerURL = 'https://www.beeradvocate.com'
            
        elif(operationChoice == 2):
            BeerCategoryTree = loadFromWorkbook(BeerCategoryTree)
            
        elif(operationChoice == 3):
            printTree(BeerCategoryTree)
            
        elif(operationChoice == 4):
            saveToWorkbook(BeerCategoryTree)
            
        elif(operationChoice == 5):
            BeerCategoryTree = scrapeCategoryItems(BeerCategoryTree, beerURL)
            
        elif(operationChoice == 6):
            BeerCategoryTree = manualKeyReassignment(BeerCategoryTree)
            
        elif(operationChoice == 7):
            BeerCategoryTree = wordBankBuilder(BeerCategoryTree)
            
        elif(operationChoice == 8):
            displayMachineLearningOptions(BeerCategoryTree)
            
        elif(operationChoice == 9):
            testOption(BeerCategoryTree)
            
            
    #    elif(operationChoice == 7):
        
        operationChoice = treeMenuOptions()  
       
        
main()
