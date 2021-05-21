''' Beer_Data_Collection

Module designed to collect beer information from BeerAdvocate.com
and store the information in xlsx files for ease of readability.


The program uses Selenium to open BeerAdvocate.com, scrape data,
and stores them in BeerCategory and Beer objects.  If there is a
disconnect in the network during the collection process, Selenium
will try reconnecting every 30 seconds until a connection is
re-established.  BeautifulSoup is used to search the html format
for tags of interest to collect the data.


The program can compile word counts for further analysis.
The data is then saved in xlsx format for each style for ease of
readability, and save all beer information in one csv format.

'''

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import BeerCategoryClass
import BeerClass
import numpy as np
import os
import re
import time

#******************************************************************************************************************************
# CONSTANTS
# the starting row of beer information
BEER_REVIEW_START_ROW = 21
BEER_FEATURES_START_ROW = 50
BEER_WORD_COUNT_START_ROW = 70

# the starting row of category information
CATEGORY_FEATURES_START_ROW = 15
CATEGORY_BEER_REVIEW_COUNT = 41

# number of features in a beer
MAX_NUMBER_OF_FEATURES = 11

# File Names:
BEER_ALL_INFO = 'Beer_All_Info\\'
NEW_BEER_ALL_INFO = 'New_Beer_All_Info\\'
FILE_DIRECTORY = 'D:\\Python Projects\\Beer Recommender Project\\'
OMITTED_WORDS = 'omitted words.txt'
KEYWORD_BANK = 'Keyword Bank\\'
SCRAPE = 'Scrape\\'

#******************************************************************************************************************************
# funtion where selenium gathers html from each web page
def seleniumGetsHTML(site):
    ''' Selenium opens the website and BeautifulSoup collects the html

    Parameters
    ----------
    site : str
        url to get html for

    Returns
    ----------
    html : bs4.BeautifulSoup data structure.
        html data structure of the url

    '''
    

    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.get(site)
    time.sleep(3)  ### Timer to allow time for the compiler to grab html

    html = BeautifulSoup(driver.page_source, 'html.parser')

    driver.close()
    driver.quit()

    return html

#******************************************************************************************************************************
## returns the text of the starting node.
# add the first node to the tree here...
def buildTree(html, site = 'https://www.beeradvocate.com/beer/styles'):
    ''' creates a tree of BeerCategory objects
    based on the layout of BeerAdvocate.com/beer/styles

    Parameters
    ----------
    html : bs4.BeautifulSoup data structure
        html structure to create tree from
        
    site : str
        url to build tree from
        defaults to "https://www.beeradvocate.com/beer/styles"

    Returns
    ----------
    html : bs4.BeautifulSoup data structure.
        returns the html for the url being opened
    
'''
    
    # create a root node
    rootNodeOfTree = BeerCategoryClass.BeerCategory()   
    key = 0
    # give root node basic attributes
    rootNodeOfTree.setCategoryName('All Beers')
    rootNodeOfTree.setCategoryKey(key)              # root node will have a key of 0 every time.
    rootNodeOfTree.setCategoryParent(None)          # root node has no parents
    rootNodeOfTree.setCategoryParentKey(key - 1)    # root node does not have a parent, thus root node's parent key is set to -1
    rootNodeOfTree.setCategory_href('/beer/styles') # since we are starting that this page manually, we will go ahead and enter this in manually
    rootNodeOfTree.setSubCategoriesExist(True)
    
    styleBreakClasses = html.findAll('div', {'class': 'stylebreak'})  # this will get all of the div tags with class 'stylebreak'
    for styleBreak in styleBreakClasses:
        key += 1
        currentCategory = BeerCategoryClass.BeerCategory()
        currentCategory.setCategoryName(styleBreak.find('b').get_text())
        currentCategory.setCategoryKey(key)
        currentCategory.setCategoryParent(rootNodeOfTree)
        currentCategory.setCategoryParentKey(rootNodeOfTree.getCategoryKey())
        currentCategory.setSubCategoriesExist(True)
        subCategoryLists = styleBreak.findAll('a')
        for subCategory in subCategoryLists:
            key += 1
            currentSubCategory = BeerCategoryClass.BeerCategory()
            currentSubCategory.setCategoryName(subCategory.get_text())
            currentSubCategory.setCategoryKey(key)
            currentSubCategory.setCategoryParent(currentCategory)
            currentSubCategory.setCategoryParentKey(currentCategory.getCategoryKey())
            currentSubCategory.setCategory_href(subCategory['href'])
            currentCategory.addSubCategory(currentSubCategory)
        rootNodeOfTree.addSubCategory(currentCategory)
    return rootNodeOfTree

    
#******************************************************************************************************************************
# Outputting tree to screen
# prints a visual representation of the categories and sub categories tree.

def printCategoryTree(currentCategory, level = 1):
    ''' Print the BeerCategory tree and its unique key to screen 

    Parameters
    ----------
    currentCategory : BeerCategory object, root of the tree
        object that holds data of beer category
        
    level : int
        defines which level of the tree is being printed
        defaults to 1

    '''

    if currentCategory.doSubCategoriesExist() == True:
        for sub in currentCategory.getSubCategories():
            print((' |    ')*level)
            print((' |    ')*level)
            print((' |    ')*(level-1) + ' |--' + sub.getCategoryName() + '  Key: ' + str(sub.getCategoryKey()))
            printCategoryTree(sub, (level+1))
        level -= 1    

#*******************************************************************************************************
# Sending tree information to an excel sheet.

def createWorkbookForEachCategories(root):
    ''' Creates xlsx workbooks to store collected information for each BeerCategory object with Beer objects

    Parameters
    ----------
    root : BeerCategory object, root of the tree
        root object of the beer category tree
        
    '''

    for category in root.getSubCategories():
        for subCategory in category.getSubCategories():
            print('Saving Style to \\New_All_Beer_Info folder: ' + subCategory.getCategoryName())
            wb = Workbook()
            categoryName = re.sub('\/', 'and', subCategory.getCategoryName())
            __addStyleToNewWorkbook(wb, subCategory, 'None', -1, index = 0)
            openFile = NEW_BEER_ALL_INFO + categoryName + '.xlsx'
            wb.save(openFile)


def __addStyleToNewWorkbook(wb, currentCategory, pName, pkey, index):
    # enters data stored in category object into xlsx file.
    # first page is information pertaining to the category.
    
    currentRow = 0
    currentColumn = 0

    categoryName = re.sub('\/', 'and', currentCategory.getCategoryName())
    wb.create_sheet(index = index, title = categoryName)
    wb.active = index
    sheet = wb.active
    
    catName = sheet.cell(row = 1, column = 1)
    catName.value = 'Category Name:'
    catName = sheet.cell(row = 1, column = 2)    
    catName.value = currentCategory.getCategoryName()

    catKey = sheet.cell(row = 2, column = 1)
    catKey.value = 'Category Key:'
    catKey = sheet.cell(row = 2, column = 2)    
    catKey.value = currentCategory.getCategoryKey()

    catParent = sheet.cell(row = 3, column = 1)
    catParent.value = 'Category Parent:'
    catParent = sheet.cell(row = 3, column = 2)    
    catParent.value = currentCategory.getCategoryParent().getCategoryName()
    
    catParentKey = sheet.cell(row = 4, column = 1)
    catParentKey.value = 'Category Parent Key:'
    catParentKey = sheet.cell(row = 4, column = 2)    
    catParentKey.value = currentCategory.getCategoryParentKey()
        
    catHREF = sheet.cell(row = 5, column = 1)
    catHREF.value = 'Category href:'
    catHREF = sheet.cell(row = 5, column = 2)
    catHREF.value = currentCategory.getCategory_href()

    catDesc = sheet.cell(row = 6, column = 1)
    catDesc.value = 'Category Description:'
    catDesc = sheet.cell(row = 6, column = 2)    
    catDesc.value = currentCategory.getCategoryDescription()

    catDesc = sheet.cell(row = 7, column = 1)
    catDesc.value = 'Minimum ABV:'
    catDesc = sheet.cell(row = 7, column = 2)    
    catDesc.value = currentCategory.getCategoryMinABV()

    catDesc = sheet.cell(row = 8, column = 1)
    catDesc.value = 'Maximum ABV:'
    catDesc = sheet.cell(row = 8, column = 2)    
    catDesc.value = currentCategory.getCategoryMaxABV()

    catDesc = sheet.cell(row = 9, column = 1)
    catDesc.value = 'Minimum IBU:'
    catDesc = sheet.cell(row = 9, column = 2)    
    catDesc.value = currentCategory.getCategoryMinIBU()

    catDesc = sheet.cell(row = 10, column = 1)
    catDesc.value = 'Maximum IBU:'
    catDesc = sheet.cell(row = 10, column = 2)    
    catDesc.value = currentCategory.getCategoryMaxIBU()

    ## category's features go here.
    currentRow = CATEGORY_FEATURES_START_ROW
    if len(currentCategory.getCategoryFeaturesMatrix()) == MAX_NUMBER_OF_FEATURES:
        feature = currentCategory.getCategoryFeaturesMatrix()
        i = 0

        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Astringency'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Body'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Alcohol'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Bitter'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Sweet'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Sour'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Salty'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Fruits'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Hoppy'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Spices'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1
        
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Malty'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = feature[i]
        currentRow += 1
        i += 1

    # save a list of beer and their review count,
    # will use this to determine how impacting the reviews will be on features
    var = sheet.cell(row = 40, column  = 1)
    var.value = 'Beer Name'
    var = sheet.cell(row = 40, column  = 2)
    var.value = 'Review Count'    
    currentRow = CATEGORY_BEER_REVIEW_COUNT
    currentColumn = 1    
    for beer in currentCategory.getCategoryBeers():
        var = sheet.cell(row = currentRow, column  = 1)
        var.value = beer.getBeerName()
        var = sheet.cell(row = currentRow, column  = 2)
        var.value = len(beer.getBeerReviewsFullContent())
        currentRow += 1

### START BEER INFORMATION
    # enter information for each beer into their own sheets after category sheet
    for eachItem in currentCategory.getCategoryBeers():
        index += 1
        beerName = eachItem.getBeerName()
        beerName = re.sub('[^A-Za-z0-9]+', '', beerName)
        wb.create_sheet(index = index, title = beerName)
        wb.active = index
        sheet = wb.active
        
        label = sheet.cell(row = 1, column = 1)
        label.value = 'Beer Name'
        var = sheet.cell(row = 1, column = 2)
        var.value = eachItem.getBeerName()
        
        label = sheet.cell(row = 2, column = 1)
        label.value = 'Beer key'
        var = sheet.cell(row = 2, column = 2)
        var.value = eachItem.getBeerKey()
        
        label = sheet.cell(row = 3, column = 1)
        label.value = 'Beer Style'
        var = sheet.cell(row = 3, column = 2)
        var.value = eachItem.getBeerStyle()
        
        label = sheet.cell(row = 4, column = 1)
        label.value = 'Beer Style Key'
        var = sheet.cell(row = 4, column = 2)
        var.value = eachItem.getBeerCategoryKey()
        
        label = sheet.cell(row = 5, column = 1)
        label.value = 'Brewery'
        var = sheet.cell(row = 5, column = 2)
        var.value = eachItem.getBeerBrewery()
        
        label = sheet.cell(row = 6, column = 1)
        label.value = 'Ave Rating'
        var = sheet.cell(row = 6, column = 2)
        var.value = eachItem.getBeerAverageRating()
        
        label = sheet.cell(row = 7, column = 1)
        label.value = 'Beer ABV'
        var = sheet.cell(row = 7, column = 2)
        var.value = eachItem.getBeerABV()
        
        label = sheet.cell(row = 8, column = 1)
        label.value = 'Beer Min IBU'
        var = sheet.cell(row = 8, column = 2)
        var.value = eachItem.getBeerMinIBU()
               
        label = sheet.cell(row = 9, column = 1)
        label.value = 'Beer Max IBU'
        var = sheet.cell(row = 9, column = 2)
        var.value = eachItem.getBeerMaxIBU()
        
        label = sheet.cell(row = 10, column = 1)
        label.value = 'Beer Description'
        var = sheet.cell(row = 10, column = 2)
        try:
            var.value = eachItem.getBeerDescription()
        except:
            var.value = 'error entering this description'

### BEER FULL REVIEWS HERE
        label = sheet.cell(row = 20, column = 1)
        label.value = 'Reviews:'
        currentRow = BEER_REVIEW_START_ROW
        currentColumn = 1
        columnWidth = 100
        for eachReview in eachItem.getBeerReviewsFullContent():
            var = sheet.cell(row = currentRow, column = currentColumn)            
            try:
                var.value = eachReview
            except:
                var.value = 'error entering this review'
            currentRow += 1

### BEER FEATURES MATRIX HERE
        # this is where we will save our beer features values
        currentRow = BEER_FEATURES_START_ROW
        if len(eachItem.getBeerFeaturesMatrix()) == MAX_NUMBER_OF_FEATURES:
            feature = eachItem.getBeerFeaturesMatrix()
            i = 0

            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Astringency'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Body'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Alcohol'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Bitter'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Sweet'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Sour'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Salty'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Fruits'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Hoppy'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Spices'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
            
            var = sheet.cell(row = currentRow, column = 1)
            var.value = 'Malty'
            var = sheet.cell(row = currentRow, column = 2)
            var.value = feature[i]
            currentRow += 1
            i += 1
### BEER WORD COUNTS START HERE
        # word count for each beer from description and reviews saved here
        currentRow = BEER_WORD_COUNT_START_ROW
        var = sheet.cell(row = currentRow, column = 1)
        var.value = 'Words'
        var = sheet.cell(row = currentRow, column = 2)
        var.value = 'Counts'
        
        currentRow += 1
        currentColumn = 1
        if eachItem.getBeerWordCount() != None:
            for keyVal in eachItem.getBeerWordCount():
                currentCell = sheet.cell(row = currentRow, column = currentColumn)
                # sends word to excel file
                variable = currentCell
                variable.value = keyVal[0]
                currentColumn += 1
                currentCell = sheet.cell(row = currentRow, column = currentColumn)
                # sends word count to excel file
                variable = currentCell
                variable.value = keyVal[1]
                currentRow += 1
                currentColumn = 1
        

# Look for any empty or blank pages and remove them from the workbook (this usually occurs at the end of the workbook)
    for sheet in wb:
        if sheet.cell(row = 1, column = 1).value == '' or sheet.cell(row = 1, column = 1).value == None:
            wb.remove(sheet)

#*******************************************************************************************************
# Get information from excel file, given that there is an implemented tree already made and information
# just needs to be filled in

def loadSubCategories(root):
    ''' Fills an empty tree with already saved data into subcategories with beer objects

    Parameters
    ----------
    root : BeerCategory object
        root object of the beer category tree

    Returns
    ----------
    root : BeerCategory object
        root object of the beer category tree
        tree is updated with category information collected in files
        
    '''
           
    # first we need to see if the name of the categories we want to collect data from exist as files
    fileList = os.listdir(FILE_DIRECTORY + BEER_ALL_INFO)
    print('Gathering File Names..')
    excelFileList = []
    fileName = BEER_ALL_INFO
    for file in fileList:
        excelFileList.append(fileName + file)  # list of file names

    print('Matching categories to their keys..')
    for excel in excelFileList:
        keyFound = False
        wb = load_workbook(excel)
        wb.active = 0
        key = int(wb.active.cell(row = 2, column = 2).value)
        for category in root.getSubCategories():
            for subCategory in category.getSubCategories():
                if subCategory.getCategoryKey() == key:
                    print('Loading Style: ' + subCategory.getCategoryName())
                    subCategory = __gatherInformation(0, wb, subCategory)
                    subCategory.setCategoryParent(category)
                    keyFound == True
                    break
            if keyFound == True:
                break

    category = tree.getSubCategories()[0]
#    for subCategory in category.getSubCategories():
#        print(subCategory.getCategoryName())
#        print(subCategory.getCategoryMaxIBU())
                    
    print('loading complete')

    return root

def __gatherInformation(index, wb, tempCategory):
    # pull saved information from the xlsx files starting with category information
    # then collect beer information for each category.
    
    wb.active = index
    sheet = wb.active
    tempCategory.setCategoryName(sheet.cell(row = 1, column = 2).value)    
    tempCategory.setCategoryKey(int(sheet.cell(row = 2, column = 2).value))
    # we set category parent to the object being iterated after this function
    tempCategory.setCategoryParentKey(int(sheet.cell(row = 4, column = 2).value))
    tempCategory.setCategory_href(sheet.cell(row = 5, column = 2).value)
    tempCategory.setCategoryDescription(sheet.cell(row = 6, column = 2).value)
    tempCategory.setCategoryMinABV(float(sheet.cell(row = 7, column = 2).value))
    tempCategory.setCategoryMaxABV(float(sheet.cell(row = 8, column = 2).value))
    tempCategory.setCategoryMinIBU(int(sheet.cell(row = 9, column = 2).value))
    tempCategory.setCategoryMaxIBU(int(sheet.cell(row = 10, column = 2).value))
    # this is where we will load our keyword data bank for this beer
    currentRow = 15
    i = 0
    features = [] 
    index += 1

    # get each beer information here.
    while index < len(wb.worksheets):
        wb.active = index
        sheet = wb.active
        
        currentColumn = 2
        currentRow = 1
        currentCell = sheet.cell(row = currentRow, column = currentColumn)
        while currentCell.value != '' and currentCell.value != None:
            tempItem = BeerClass.Beer()

            value = currentCell.value
            tempItem.setBeerName(value)
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            
            value = currentCell.value
            if value == '' or value == None:
                value = 0
            tempItem.setBeerKey(int(value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            
            value = currentCell.value
            if value == '' or value == None:
                value = 'NA'
            tempItem.setBeerStyle(value)
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            
            value = currentCell.value
            if value == '' or value == None:
                value = 0
            tempItem.setBeerCategoryKey(int(value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            
            value = currentCell.value
            if value == '' or value == None:
                value = 'NA'
            tempItem.setBeerBrewery(value)
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            
            value = currentCell.value
            if value == '' or value == None:
                value = 0.0
            tempItem.setBeerABV(float(value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            
            value = currentCell.value
            if value == '' or value == None:
                value = 0.0
            tempItem.setBeerAverageRating(float(value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            
            value = currentCell.value
            if value == '' or value == None:
                value = 0
            tempItem.setBeerMinIBU(int(value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            
            value = currentCell.value
            if value == '' or value == None:
                value = 0
            tempItem.setBeerMaxIBU(int(value))
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            
            value = currentCell.value
            if value == '' or value == None:
                value = 'NA'
            tempItem.setBeerDescription(value)

### LOAD BEER REVIEWS
            currentRow = BEER_REVIEW_START_ROW
            currentColumn = 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            while currentCell.value != '' and currentCell.value != None:
                value = currentCell.value
                tempItem.addBeerReviewsFullContent(value)
                currentRow += 1
                currentCell = sheet.cell(row = currentRow, column = currentColumn)

### LOAD BEER FEATURES
            currentRow = BEER_FEATURES_START_ROW
            currentColumn = 2
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            features = []
            while currentCell.value != '' and currentCell.value != None:
                features.append(currentCell.value)
                currentRow += 1
                currentCell = sheet.cell(row = currentRow, column = currentColumn)
            tempItem.setBeerFeaturesMatrix(features)
            
### LOAD BEER WORD COUNTS
# UNDER CONSTRUCTION
#            currentRow += 6
#            currentColumn = 1
#            currentCell = sheet.cell()
#            while currentCell.value != '' and currentCell.value != None:


### ADD BEER TO CATEGORY
        tempCategory.addCategoryBeer(tempItem)

### GO TO NEXT PAGE
        index += 1

    return tempCategory

#*************************************************************************************************************
## these sets of functions web scrapes heb.com in each category and loads items to their respective categories
## in the category tree.

def startGetCategoryItems(root, site = 'https://www.beeradvocate.com'):
    ''' Open BeerAdvocate.com using Selenium and collects the first 50 available beers of each category

    url extensions for each category are collected during the tree building process
    extends the site variable to open that webpage using selenium

    Parameters
    ----------
    root : BeerCategory object
        root of the beer category tree

    site : str
        base url to start gathering objects process from
        defaults to "https://www.beeradvocate.com"

    Returns
    ----------
    root : BeerCategory object
        root of the beer cateogry tree
        tree is updated with list of beer objects collected from beeradvocate.com
    
    '''

    sinceEpoch = time.time()
    startTimeObj = time.localtime(sinceEpoch)
    driver = webdriver.Chrome(ChromeDriverManager().install())
    currentSite = site + root.getCategory_href()
    driver.get(currentSite)
    time.sleep(3)

    # we know that there are no items in each category in root's list of sub categories
    # we move on to each category's sub category where we will find our beer.

    # first we will see if there are any category files already in existence due to interrupts    
    key = 0
    for eachCategory in root.getSubCategories():
        for eachSubCategory in eachCategory.getSubCategories():
            # this is done to skip any category scraping for data we already have after loading
            # if we have more than one beer in this eachSubCategory, then we know that we already
            # have this information from a previous scrape session.
            # this has to be implemented because of SPECTRUM's wonderful internet connection....
            # to save time
            if len(eachSubCategory.getCategoryBeers()) == 0:
                print('Scraping ' + eachSubCategory.getCategoryName())
                key = __openSubCategoryPages(eachSubCategory, driver, site, key)
    
    driver.close()
    driver.quit()
    
    sinceEpoch = time.time()
    endTimeObj = time.localtime(sinceEpoch)
    
    print('Started: %d:%d' %(startTimeObj.tm_hour, startTimeObj.tm_min))
    print('Finished: %d:%d' %(endTimeObj.tm_hour, endTimeObj.tm_min))

    return root

def __cleanupDoubleValueStrings(string):

    print(string)

    if string != '' or string != None:
        newString = re.sub('[a-zA-Z()\s$\/%:]', '', string)
        minVal = maxVal = ''
        dividerFound = False
        for c in newString:
            if c == '-':
                dividerFound = True
                continue
            if dividerFound == False:
                minVal += c
#                print('minVal: ' + str(minVal))
            else:
                if c == '|' or c == '%':
                    break
                maxVal += c
#                print('maxVal: ' + str(maxVal))
    print('minVal: ' + str(minVal))
    print('maxVal: ' + str(maxVal))
    return minVal, maxVal

def __cleanUpSingleValueString(string):

    if string != '' or string != None:
        newString = re.sub('[a-zA-Z()\s$/%:]', '', string)
    return newString
    

def __openSubCategoryPages(currentCategory, dr, site, key):

    # open pages of sub categories to get to list of beer   
    currentSite = site + currentCategory.getCategory_href()
    dr.execute_script("window.open(''); ")
    dr.switch_to.window(dr.window_handles[1])
    dr.get(currentSite)
    time.sleep(3)

    allHTML = BeautifulSoup(dr.page_source, 'html.parser')

    # details of beer style can be found in first <div> tag of <div id="ba-content">
    detailsTag = allHTML.find('div', {'id' : 'ba-content'})
    detailsTag = detailsTag.findAll('div')[0]
    currentCategory.setCategoryDescription(detailsTag.get_text())
    decimals = detailsTag.findAll('span')
    minABV, maxABV = __cleanupDoubleValueStrings(decimals[0].get_text())
    minIBU, maxIBU = __cleanupDoubleValueStrings(decimals[1].get_text())
    try:
        currentCategory.setCategoryMinABV(float(minABV))
        currentCategory.setCategoryMaxABV(float(maxABV))
    except:
        print('had an issue getting ABV in ' + currentCategory.getCategoryName())
#        currentCategory.setCategoryManualEditFlag(True)
    try:
        currentCategory.setCategoryMinIBU(float(minIBU))
        currentCategory.setCategoryMaxIBU(float(maxIBU))
    except:
        print('had an issue getting IBU in ' + currentCategory.getCategoryName())
#        currentCategory.setCategoryManualEditFlag(True)
    
    # only one tbody tag that holds line items of beer
    tbodyTag = allHTML.find('tbody')
    trTags = tbodyTag.findAll('tr')
    # the first three <tr> tags are as follows
    # tag 1: non-important
    # tag 2: links to the next 50 beers, not important yet.
    # tag 3: table headers allowing for re-ordering of beers, not very important
    # tag 4: start of beers, very important.
    # ...
    # ...
    # tag last: links to get to next 50 beers, very important.
    # starting at index 3, we will get the beer information available on this page.

    # this is where we go to each beer's individual website.
    
    index = 3 
    maxCount = len(trTags) - 2
    while index < maxCount:
        key += 1
        current_href = trTags[index].find('a')['href']
        # open pages of sub categories to get to list of beer   
        currentSite = site + current_href
        dr.execute_script("window.open(''); ")
        dr.switch_to.window(dr.window_handles[2])
        dr.get(currentSite)
        time.sleep(3)
        thisBeerHTML = BeautifulSoup(dr.page_source, 'html.parser')
        thisBeer = BeerClass.Beer()
        # since beerName is name of beer and name of brewery on the same string,
        # we get the the breweryName and sub out of beerName
        beerName = thisBeerHTML.find('div', {'class': 'titleBar'})
        breweryName = beerName.find('span').get_text()
        beerName = beerName.get_text()
        beerName = re.sub(breweryName, '', beerName)
        thisBeer.setBeerName(beerName)
        thisBeer.setBeerBrewery(breweryName)
        thisBeer.setBeerKey(key)
        thisBeer.setBeerCategoryKey(currentCategory.getCategoryKey())
        thisBeer.setBeerStyle(currentCategory.getCategoryName())
        # for finding beer stats
        beerStats = thisBeerHTML.findAll('dd', {'class': 'beerstats'})
        abv = __cleanUpSingleValueString(beerStats[1].get_text())
        try:
            thisBeer.setBeerABV(float(abv))
        except:
            print(thisBeer.getBeerName() + ' does not have an ABV on the website.\nSubstituting with the category average ABV')
            try:
                thisBeer.setBeerABV((float(minABV) + float(maxABV)) / 2)
            except:
                thisBeer.setBeerManualEditFlag(True)
        thisBeer.setBeerAverageRating(float(beerStats[3].find('span', {'class': 'ba-ravg Tooltip'}).get_text()))
            
        # for finding notes
        notes = thisBeerHTML.find('div', {'style': 'clear:both; margin:0; padding:0px 20px; font-size:1.05em;'})
        thisBeer.setBeerDescription(notes.get_text())
        try:
            thisBeer.setBeerMinIBU(float(minIBU))
            thisBeer.setBeerMaxIBU(float(maxIBU))
        except:
            thisBeer.setBeerManualEditFlag(True)
        # for finding all beer reviews
        beerReviews = thisBeerHTML.findAll('div', {'id': 'rating_fullview_content_2'})
        for each in beerReviews:
            thisBeer.addBeerReviewsFullContent(each.get_text())
        
        currentCategory.addCategoryBeer(thisBeer)
    
        dr.close()
        dr.switch_to.window(dr.window_handles[1])
        index += 1
    
    dr.close()
    dr.switch_to.window(dr.window_handles[0])

    # save to a new excel document after each category is scraped
    # because SPECTRUM...
    wb = Workbook()
    print('Saving current category to \\Scrape folder: ' + currentCategory.getCategoryName())
    __addStyleToNewWorkbook(wb, currentCategory, currentCategory.getCategoryName(), currentCategory.getCategoryKey(), index = 0)
    name = re.sub('\/', 'and',currentCategory.getCategoryName())
    openFile = SCRAPE + name + '.xlsx'
    wb.save(openFile)


    return key  # return the value of the current key to avoid duplicate key values..
                # consider changing the format of the key value to avoid this return..

#*******************************************************************************************************************************
# Manual Key Reassignment to individual beers so that each beer is garaunteed a unique key
# This is to be used in the case that the internet connection drops and I has to pick up where I left off.

def reassignKeys(root):
    ''' Reassign key values to all beer objects.
    In the case that the internet connection is disrupted,
    the user can redefine beer keys so that each key is definitively unique

    Parameters
    ----------
    root : BeerCategory object
        root of the beer category tree

    Returns
    ----------
    root : BeerCategory object
        root of the beer category tree
        tree is updated with unique keys for each beer object

    
    '''
    key = 0
    for category in root.getSubCategories():
        for subCategory in category.getSubCategories():
            for beer in subCategory.getCategoryBeers():
                key += 1
                beer.setBeerKey(key)
    return root

#**********************************************************************************************************************************
# set of functions build keyword data banks

def compileWordCounts(root):
    ''' Returns a dictionary of words and word counts extracted from collected reviews.
    This exludes a short list of unwanted words such as "and, a, the, etc..".
    The words kept comes from an approved list of words that define a beer's characteristic

    Parameters
    ----------
    root : BeerCategory object
        root of the beer category tree

    Returns
    ----------
    root : BeerCategory object
        root of the beer category tree
        updated with compiled word counts for each beer object of each class
    
    '''

    # remove unwanted words.
    invalidWords = []
    file = open(KEYWORD_BANK + OMITTED_WORDS, 'r')
    for word in file:
        word = re.sub('\n', '', word)
        invalidWords.append(word)
        
    # dictionary for all of the words in the category.
    # it gets broken down into sub categories
#    categoryWordCountDictionary = {}     # will house dictionaries in category
    subCategoryWordCountDictionary = {}  # will house dictionaries in sub category
#    currentBeerWordCountDictionary = {}
    
    # for getting category word dictionaries of words
    for category in root.getSubCategories():        
        # for getting sub category dictionaries of words
        for subCategory in category.getSubCategories():
#            subCategoryWordCountDictionary = {}  # will house dictionaries in sub category
            # gets words from the description of the sub category
            # as stated by BeerAdvocate                    
            # for getting beer dictionaries of words
            
            for beer in subCategory.getCategoryBeers():                
                currentBeerWordCountDictionary = {} # empty dictionary of word count for current beer
                # gets words from the description of the beer
                # as stated by the brewery
                allWords = beer.getBeerDescription()
                allWords = re.sub('[^A-Za-z]+', ' ', allWords)
                allWords = allWords.lower()
                currentWord = ''
                for c in allWords:
                    if c == ' ' and currentWord != '':
                        
                        isValid = True
                        for invalid in invalidWords:
                            if currentWord == invalid:
                                isValid = False
                                break
                            
                        if isValid == True:
                            if currentWord in currentBeerWordCountDictionary:
                                count = currentBeerWordCountDictionary[currentWord]
                                count += 1
                                currentBeerWordCountDictionary[currentWord] = count
                            else:
                                currentBeerWordCountDictionary[currentWord] = 1
                        currentWord = ''
                        
                    else:
                        currentWord += c
                        
                # then we get the words from all of the reviews
                # left by users on BeerAdvocate
                for review in beer.getBeerReviewsFullContent():
                    allWords = review
                    allWords = re.sub('[^A-Za-z]+', ' ', allWords)
                    allWords = allWords.lower()
                    currentWord = ''
                    for c in allWords:
                        if c == ' ' and currentWord != '':
                            
                            isValid = True
                            for invalid in invalidWords:
                                if currentWord == invalid:
                                    isValid = False
                                    break

                            if isValid == True:
                                if currentWord in currentBeerWordCountDictionary:
                                    count = currentBeerWordCountDictionary[currentWord]
                                    count += 1
                                    currentBeerWordCountDictionary[currentWord] = count
                                else:

                                    currentBeerWordCountDictionary[currentWord] = 1
                            currentWord = ''
                        else:
                            currentWord += c
                currentBeerWordCountDictionary = sorted(currentBeerWordCountDictionary.items(), key=lambda x: x[1], reverse=True)
                beer.setBeerWordCount(currentBeerWordCountDictionary)

            # here we will add all of the beer words to the sub category words so that we can build a profile of the categories as well
            for word in currentBeerWordCountDictionary:
                if word[0] in subCategoryWordCountDictionary:
                    count = subCategoryWordCountDictionary[word[0]]
                    count += word[1]
                    subCategoryWordCountDictionary[word[0]] = count
                else:
                    subCategoryWordCountDictionary[word[0]] = word[1]
            subCategory.setCategoryFeaturesMatrix(subCategoryWordCountDictionary)
    return root


#*************************************************************************************************************************************
# function to get a combined word count from all beers.  This is to assist with sifting for beer related words manually

def combineWordCounts(root):
    ''' compiles a list of all accepted words and respective word counts after compileWordCount is complete.
    Saves the list of all words and word counts to an xlsx file.
    This was used to determine the frequency of all accepted words in all reviews collected,
    and to build a key word bank that defined the features of a beer.
    
    Parameters
    ----------
    root : BeerClass object
        root of the beer category tree

    '''
    combineWordCount = {}

    for category in root.getSubCategories():
        for subCategory in category.getSubCategories():
            for beer in subCategory.getCategoryBeers():
                for wordCount in beer.getBeerWordCount():
                    if wordCount[0] in combineWordCount:
                        count = combineWordCount[wordCount[0]]
                        count += wordCount[1]
                        combineWordCount[wordCount[0]] = count
                    else:
                        combineWordCount[wordCount[0]] = wordCount[1]
    combineWordCount = sorted(combineWordCount.items(), key=lambda x: x[1], reverse=True)    
    wb = Workbook()
    categoryName = 'Combine Word Count'
    fileName = KEYWORD_BANK
    openFile = fileName + categoryName + '.xlsx'
    wb.active = 0
    sheet = wb.active
    cell = sheet.cell(row = 1, column = 1)
    cell.value = 'Words'
    cell = sheet.cell(row = 1, column = 2)
    cell.value = 'Counts'
    currentRow = 2
    for wordCount in combineWordCount:
        cell = sheet.cell(row = currentRow, column = 1)
        cell.value = wordCount[0]
        cell = sheet.cell(row = currentRow, column = 2)
        cell.value = wordCount[1]
        currentRow += 1
    
    wb.save(openFile)

#***********************************************************************************************************************************
# function to build a features matrix for each beer by comparing word counts with those in the keyword data bank
# each count on a word will impact the order of magnitude for the feature it is stored under

def compileFeaturesDefinitions():
    ''' returns a dictionary of beer feature and their respective word key and impact value.
    All word's impact value are currently set to 1
    Open and load feature definitions
    This determines which feature of a beer each word counted in a beer's review impacts
    This is essentially a feature definition file being loaded.
    
    Returns
    ----------
    features : dict
        dictionary of features and their definitions via word in review
        features = { feature : { word1 : impact1 } ... { wordN : impactN } }
    
    '''
    features = {}    # an array of 19 features, dictionary of word with a magnitude of impact
    
    fileName = KEYWORD_BANK
    file = 'Beer Descriptors Simplified.xlsx'
    wb = load_workbook(fileName + file)

    # first we need to load the beer descriptors (features) and their respective impact magnitudes
    index = 0
    wb.active = index
    sheet = wb.active
    currentRow = 1
    currentColumn = 1
    currentCell = sheet.cell(row = currentRow, column = currentColumn)
    while index < 3:
        currentFeature = {}
        featureName = currentCell.value
        currentRow += 1
        currentCell = sheet.cell(row = currentRow, column = currentColumn)
        while currentCell.value != '' and currentCell.value != None:
            
            feature = currentCell.value
            currentColumn += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            impact = int(currentCell.value)
            currentFeature.update({feature : impact})
            
            currentRow += 1
            currentColumn -= 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)

        features.update({featureName : currentFeature})
            
        currentRow = 1
        currentColumn += 2
        currentCell = sheet.cell(row = currentRow, column = currentColumn)

        # if row 2 of the next feature is empty, we know we need to move to the next sheet.
        # and reset the currentCell to 2, 1
        if currentCell.value == '' or currentCell.value == None:
            index += 1
            if index == 3:
                break
            wb.active = index
            sheet = wb.active
            currentRow = 1
            currentColumn = 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)

    # next we iterate through the word counts of each beer and keep the words that match those in beerDescriptors.
    return features

#******************************************************************************************************************************
# This function cycles through each feature's list of keywords (definition of features if you will)
# If a word from the beer's word count matches the feature's keyword, then it adds plus 1 to the feature matrix for that feature
# matrix: [astringent, mouthfeel, alcohol, bitter, sweet, sour, salty, fruity, hoppy, spicy, malty]
def wordCountToFeatures(root, features):
    ''' Returns the beer style/category tree after populating beer features matrix for each beer
    Convert the word counts into beer features based on the predefined key word bank.
    Key word bank simply states which word found impacts which feature in a beer.
    For example, the word "dry" appearing 12 times from a beer's collected reviews will add 12 to the "Astringency" feature.

    Parameters
    ----------
    root : BeerCategory object
        root of the beer category tree

    features : dict
        dictionary of features and their definitions via word in review
        features = { feature : { word1 : impact1 } ... { wordN : impactN } }

    Returns
    ----------
    root : BeerCategory object
        root of the beer category tree    
    
    '''
    for category in root.getSubCategories():
        for subCategory in category.getSubCategories():
            for beer in subCategory.getCategoryBeers():   
                matrixOfBeerFeatures = []           # n x 1 matrix
                for feature, impact in features.items():
                    magnitudeOfCurrentFeature = 0
                    for word in beer.getBeerWordCount():
                        for key in impact:
                            if word[0] == key:
                                magnitudeOfCurrentFeature += (int(word[1]) * int(impact[key]))
                    matrixOfBeerFeatures.append(magnitudeOfCurrentFeature)
                beer.setBeerFeaturesMatrix(matrixOfBeerFeatures)

    # here we calculate the features matrix for sub categories or styles
    # we sum up the features matrices of all beer in the sub category and divide by the number of beers in that sub category or style
    # we find the mean features matrix that represents the sub category or style.

    for category in root.getSubCategories():
        for subCategory in category.getSubCategories():
            matrixOfSubCategoryFeatures = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
            for beer in subCategory.getCategoryBeers():
                matrixOfSubCategoryFeatures = np.add(matrixOfSubCategoryFeatures, beer.getBeerFeaturesMatrix())
            matrixOfSubCategoryFeatures = matrixOfSubCategoryFeatures / len(subCategory.getCategoryBeers())
            subCategory.setCategoryFeaturesMatrix(matrixOfSubCategoryFeatures)
    
    return root

#***************************************************************************************************************************************
# Functions to gather data on select user inputs and store the data in a file.

def gatherUserInputInformation():
    ''' Collect information from BeerAdvocate.com on previously uncollected beer to use as user input,
    and saves the information to an xlsx file.
    Currently, this is a static list of beer being used as input to test the classification recommendation systems.
        
    '''
    
    startPage = 'https://www.beeradvocate.com/'
    beerURLs = ['beer/profile/148/11436/',      # Scottish Ale 80 Shillings
                'beer/profile/335/2904/',       # IPA American Mad Hatter
                'beer/profile/199/178740/',     # IPA American Sculpin Pineapple
                'beer/profile/31987/194673/',   # Stout Oatmeal Black House Nitro
                'beer/profile/137/30936/',      # Fruit and Field Apricot
                'beer/profile/3760/9781/',      # Porter American Coffee Porter
                'beer/profile/179/71896/',      # Sour - Flanders Oud Bruin Le Serpent Cerise
                'beer/profile/192/275731/',     # Wheat Beer Witbier Fat Tire Belgian White
                'beer/profile/8/46767/',        # Pale Ale American Drifter
                'beer/profile/26762/76498/']    # Porter Imperial Hellfighter

    dummyParent = BeerCategoryClass.BeerCategory()
    userCategory = BeerCategoryClass.BeerCategory()

    dummyParent.setCategoryName('')
    dummyParent.setCategoryKey(-1)
    
    userCategory.setCategoryName('User Input')
    userCategory.setCategoryKey(100)
    userCategory.setCategoryParent(dummyParent)
    userCategory.setCategoryParentKey(-1)

    key = 9000

    for currentBeerURL in beerURLs:
        currentSite = startPage + currentBeerURL
        thisBeerHTML = seleniumGetsHTML(currentSite)
        key += 1
        
        thisBeer = BeerClass.Beer()
        # since beerName is name of beer and name of brewery on the same string,
        # we get the the breweryName and sub out of beerName
        beerName = thisBeerHTML.find('div', {'class': 'titleBar'})
        breweryName = beerName.find('span').get_text()
        beerName = beerName.get_text()
        beerName = re.sub(breweryName, '', beerName)
        thisBeer.setBeerName(beerName)
        thisBeer.setBeerBrewery(breweryName)
        thisBeer.setBeerKey(key)
        # for finding beer stats
        beerStats = thisBeerHTML.findAll('dd', {'class': 'beerstats'})
        beerStyle = beerStats[0].find('a').get_text()
        thisBeer.setBeerStyle(beerStyle)

        # to get the category key, we will navigate through all of the styles and find the name that matches the name of this beer's style.

        excelList = os.listdir(FILE_DIRECTORY + BEER_ALL_INFO)
        desiredFile = ''
        for excel in excelList:
            if excel == beerStyle:
                desiredFile = BEER_ALL_INFO + file + '.xlsx'
                wb = load_workbook(desiredFile)
                wb.active = 1
                sheet = wb.active
                thisBeer.setBeerCategoryKey(sheet.cell(row = 2, column = 2).value) # find a way to fix this
                break
        

        
        abv = cleanupSingleDecimalStrings(beerStats[1].get_text())
        try:
            thisBeer.setBeerABV(float(abv))
        except:
            print(thisBeer.getBeerName() + ' does not have an ABV on the website.\nSubstituting with the category average ABV')
            try:
                thisBeer.setBeerABV((float(minABV) + float(maxABV)) / 2)
            except:
                thisBeer.setBeerManualEditFlag(True)
        thisBeer.setBeerAverageRating(float(beerStats[3].find('span', {'class': 'ba-ravg Tooltip'}).get_text()))
            
        # for finding notes
        notes = thisBeerHTML.find('div', {'style': 'clear:both; margin:0; padding:0px 20px; font-size:1.05em;'})
        thisBeer.setBeerDescription(notes.get_text())
        try:
            thisBeer.setBeerMinIBU(float(minIBU))
            thisBeer.setBeerMaxIBU(float(maxIBU))
        except:
            thisBeer.setBeerManualEditFlag(True)
        # for finding all beer reviews
        beerReviews = thisBeerHTML.findAll('div', {'id': 'rating_fullview_content_2'})
        for each in beerReviews:
            thisBeer.addBeerReviewsFullContent(each.get_text())
        
        userCategory.addCategoryBeer(thisBeer)

        
    invalidWords = []
    file = open(KEYWORD_BANK + OMITTED_WORDS, 'r')
    for word in file:
        word = re.sub('\n', '', word)
        invalidWords.append(word)
        
    for beer in userCategory.getCategoryBeers():                
        currentBeerWordCountDictionary = {} # empty dictionary of word count for current beer
        # gets words from the description of the beer
        # as stated by the brewery
        allWords = beer.getBeerDescription()
        allWords = re.sub('[^A-Za-z]+', ' ', allWords)
        allWords = allWords.lower()
        currentWord = ''
        for c in allWords:
            if c == ' ' and currentWord != '':
                            
                isValid = True
                for invalid in invalidWords:
                    if currentWord == invalid:
                        isValid = False
                        break
                                
                if isValid == True:
                    if currentWord in currentBeerWordCountDictionary:
                        count = currentBeerWordCountDictionary[currentWord]
                        count += 1
                        currentBeerWordCountDictionary[currentWord] = count
                    else:
                        currentBeerWordCountDictionary[currentWord] = 1
                currentWord = ''
                            
            else:
                currentWord += c
                            
        # then we get the words from all of the reviews
        # left by users on BeerAdvocate
        for review in beer.getBeerReviewsFullContent():
            allWords = review
            allWords = re.sub('[^A-Za-z]+', ' ', allWords)
            allWords = allWords.lower()
            currentWord = ''
            for c in allWords:
                if c == ' ' and currentWord != '':
                                
                    isValid = True
                    for invalid in invalidWords:
                        if currentWord == invalid:
                            isValid = False
                            break

                    if isValid == True:
                        if currentWord in currentBeerWordCountDictionary:
                            count = currentBeerWordCountDictionary[currentWord]
                            count += 1
                            currentBeerWordCountDictionary[currentWord] = count
                        else:

                            currentBeerWordCountDictionary[currentWord] = 1
                    currentWord = ''
                else:
                    currentWord += c
        currentBeerWordCountDictionary = sorted(currentBeerWordCountDictionary.items(), key=lambda x: x[1], reverse=True)
        beer.setBeerWordCount(currentBeerWordCountDictionary)

                
    beerFeatures = compileFeaturesDefinitions()
    
    for beer in userCategory.getCategoryBeers():   
        matrixOfBeerFeatures = []           # n x 1
        for feature, impact in beerFeatures.items():
            magnitudeOfCurrentFeature = 0
            for word in beer.getBeerWordCount():
                for key in impact:
                    if word[0] == key:
                        magnitudeOfCurrentFeature += (int(word[1]) * int(impact[key]))
            matrixOfBeerFeatures.append(magnitudeOfCurrentFeature)
        beer.setBeerFeaturesMatrix(matrixOfBeerFeatures)

    wb = Workbook()
    __addStyleToNewWorkbook(wb, userCategory, userCategory.getCategoryName(), userCategory.getCategoryKey(), index = 0)
    name = re.sub('\/', 'and',userCategory.getCategoryName())
    file = SCRAPE
    openFile = file + name + '.xlsx'
    wb.save(openFile)



#********************************************************************************************************************************
# this function loads the basic information for all 5700 beer for machine learning application
# with the option to load user beer input from file as well.
def loadBeerInformation(getCategoryDictionary = False):
    ''' Returns a list of beer objects loaded from xlsx style/category files

    Parameters
    ----------
    getCategoryDictionary : bool
        determines of function returns category dictionary
        currently does not do anything during the process
        defaults to False

    Returns
    ----------
    beerList : list
        list of beer objects

    '''
    
    beerList = []

    # first we need to see if the name of the categories we want to collect data from exist as files
    fileList = os.listdir(FILE_DIRECTORY + BEER_ALL_INFO)
    print('Gathering File Names..')
    fileName = BEER_ALL_INFO
    for file in fileList:
        wb = load_workbook(fileName + file)
        if getCategoryDictionary == True:
            index = 0
            wb.active = index
            sheet = wb.active
            dictionary[float(sheet.cell(row = 2, column = 2).value)] = str(sheet.cell(row = 1, column = 2).value)
        
        index = 1
        print('Getting beer information from ' + file)
        while index < len(wb.worksheets):
            beer = BeerClass.Beer()            
            wb.active = index
            sheet = wb.active
            
            currentRow = 1
            currentColumn = 2
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerName(re.sub('\n\0','',currentCell.value))
            
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerKey(int(currentCell.value))
            
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerStyle(re.sub('\n\0','',currentCell.value))
            
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerCategoryKey(int(currentCell.value))
            
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerBrewery(re.sub('\n\0','',currentCell.value))
            
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerAverageRating(float(currentCell.value))
            currentRow += 1
            
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerABV(float(currentCell.value))
            currentRow += 1
            
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerMinIBU(int(currentCell.value))
            currentRow += 1
            
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerMaxIBU(int(currentCell.value))
            
            currentRow += 1
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerDescription(re.sub('\n\0','',currentCell.value))

            features = []
            currentRow = BEER_FEATURES_START_ROW
            currentCell = sheet.cell(row = currentRow, column = currentColumn)
            while currentCell.value != '' and currentCell.value != None:
                features.append(int(currentCell.value))
                currentRow += 1
                currentCell = sheet.cell(row = currentRow, column = currentColumn)
            beer.setBeerFeaturesMatrix(features)
            beerList.append(beer)
            index += 1
            
    return beerList

#******************************************************************************************************************************
# Saves beer information to a csv file.

def saveBeerInformation(beers):
    ''' Saves a list of beer objects with their data in a csv format
    
    Parameters
    ----------
    beers : list
        list of beer objects    
    
    '''
    wb = Workbook()
    wb.create_sheet(index = 0)
    wb.active = 0
    sheet = wb.active
    
    label = sheet.cell(row = 1, column = 1)
    label.value = 'Name'
    label = sheet.cell(row = 1, column = 2)
    label.value = 'Key'
    label = sheet.cell(row = 1, column = 3)
    label.value = 'Style'
    label = sheet.cell(row = 1, column = 4)
    label.value = 'Style Key'
    label = sheet.cell(row = 1, column = 5)
    label.value = 'Brewery'
    label = sheet.cell(row = 1, column = 6)
    label.value = 'Description'
    label = sheet.cell(row = 1, column = 7)
    label.value = 'Ave Rating'
    label = sheet.cell(row = 1, column = 8)
    label.value = 'ABV'
    label = sheet.cell(row = 1, column = 9)
    label.value = 'Min IBU'
    label = sheet.cell(row = 1, column = 10)
    label.value = 'Max IBU'
    label = sheet.cell(row = 1, column = 11)
    label.value = 'Astringency'
    label = sheet.cell(row = 1, column = 12)
    label.value = 'Body'
    label = sheet.cell(row = 1, column = 13)
    label.value = 'Alcohol'
    label = sheet.cell(row = 1, column = 14)
    label.value = 'Bitter'
    label = sheet.cell(row = 1, column = 15)
    label.value = 'Sweet'
    label = sheet.cell(row = 1, column = 16)
    label.value = 'Sour'
    label = sheet.cell(row = 1, column = 17)
    label.value = 'Salty'
    label = sheet.cell(row = 1, column = 18)
    label.value = 'Fruits'
    label = sheet.cell(row = 1, column = 19)
    label.value = 'Hoppy'
    label = sheet.cell(row = 1, column = 20)
    label.value = 'Spices'
    label = sheet.cell(row = 1, column = 21)
    label.value = 'Malty'
    label = sheet.cell(row = 1, column = 22)
    
    currentRow = 2
    currentColumn = 1       

    for eachItem in beers:        
        
        label = sheet.cell(row = currentRow, column = currentColumn)
        label.value = re.sub(',', ' ', eachItem.getBeerName())
        currentColumn += 1
        
        label = sheet.cell(row = currentRow, column = currentColumn)
        label.value = eachItem.getBeerKey()
        currentColumn += 1
        
        label = sheet.cell(row = currentRow, column = currentColumn)
        label.value = eachItem.getBeerStyle()
        currentColumn += 1
        
        label = sheet.cell(row = currentRow, column = currentColumn)
        label.value = eachItem.getBeerCategoryKey()
        currentColumn += 1
        
        label = sheet.cell(row = currentRow, column = currentColumn)
        label.value = re.sub(',', ' ', eachItem.getBeerBrewery())
        currentColumn += 1
        
        label = sheet.cell(row = currentRow, column = currentColumn)
        label.value = re.sub(',', ' ', eachItem.getBeerDescription())
        currentColumn += 1
        
        label = sheet.cell(row = currentRow, column = currentColumn)
        label.value = eachItem.getBeerAverageRating()
        currentColumn += 1
        
        label = sheet.cell(row = currentRow, column = currentColumn)
        label.value = eachItem.getBeerABV()
        currentColumn += 1
        
        label = sheet.cell(row = currentRow, column = currentColumn)
        label.value = eachItem.getBeerMinIBU()
        currentColumn += 1
               
        label = sheet.cell(row = currentRow, column = currentColumn)
        label.value = eachItem.getBeerMaxIBU()
        currentColumn += 1

        for feature in eachItem.getBeerFeaturesMatrix():
               
            label = sheet.cell(row = currentRow, column = currentColumn)
            label.value = feature
            currentColumn += 1

#        for review in eachItem.getBeerReviewsFullContent():
#            accumulated
#            label = sheet.cell(row = currentRow, column = currentColumn)
#            label.value = eachItem.getBeerDescription()

        currentColumn += 1        

        currentRow += 1
        currentColumn = 1
    
    
    openFile = 'New_Beer_All_Info\\Beer_Data_Set.csv'
    wb.save(openFile)


#**************************************************************************************************************************************************
# main

def main():
    ''' main function, displays options to run project '''
    # our tree object where we append the nodes and their respective data values.
    BeerCategoryTree = BeerCategoryClass.BeerCategory()

    beerURL = 'https://www.beeradvocate.com/'
    
    choice = -1
    while choice < 0 or choice > 7:
        print('\nData Collection')
        print('0. Go back: this will erase all loaded data')
        print('1. Build Empty Tree: build empty tree used to scrape website for information')
        print('2. Load Tree: load information from files for all beer styles and beer into the tree')
        print('3. Print Category Tree: print the tree of beer styles and unique keys to screen')
        print('4. Save Tree: save the tree to multiple excel files')
        print('5. Scrape for Items: scrape website for items in each category and add to empty tree.\n   This also saves the data during the process incase of internet drops')
        print('6. Manual Key Reassignment: reassigns keys to each beer in saved files\n   Due to internet interruptions, multiple beers may have the same key')
        print('7. Word Count Options: display a list of options for word counts')
        print('8. Get User Input Beer: scrape information for user input beer and store in file')
        try:
            choice = int(input())
        except:
            print('\nInvalid choice, please choose wisely.')
        if(choice == 1):
            beerURL = 'https://www.beeradvocate.com'
            print('\nBuilding category tree..\n')
            # gets the html via selenium and beautifulsoup combination
            # since we are starting a '/beer/styles', we will go ahead and append that
            # here to get the html for that page.
            customStartingPoint = '/beer/styles'
            soupSource = seleniumGetsHTML(beerURL + customStartingPoint)
            # get the tree's starting node.
            BeerCategoryTree = buildTree(soupSource)
            print('\nTree has been built!')
            
        elif(choice == 2):
            # if option two is chosen and the tree has not been built, build the tree first.
            if BeerCategoryTree.getCategoryKey() == -1:
                beerURL = 'https://www.beeradvocate.com'
                print('\nBuilding category tree..\n')
                # gets the html via selenium and beautifulsoup combination
                # since we are starting a '/beer/styles', we will go ahead and append that
                # here to get the html for that page.
                customStartingPoint = '/beer/styles'
                soupSource = seleniumGetsHTML(beerURL + customStartingPoint)
                # get the tree's starting node.
                BeerCategoryTree = buildTree(soupSource, beerURL)
                print('\nTree has been built!')
                
            print('\nLoading beer information into initialized tree from workbooks..')
            BeerCategoryTree = loadSubCategories(BeerCategoryTree)
            
        elif(choice == 3):
            print('\nPrinting category tree to screen..\n')
            print(BeerCategoryTree.getCategoryName())
            printCategoryTree(BeerCategoryTree)
            print('\nTree has been printed!')
            
        elif(choice == 4):
            print('\nSaving to file..')
            createWorkbookForEachCategories(BeerCategoryTree)
            print('\nSaving complete!')
            
        elif(choice == 5):
            beerURL = 'https://www.beeradvocate.com'
            print('\nGathering item information from website ' + beerURL + '..')
            BeerCategoryTree = startGetCategoryItems(BeerCategoryTree, beerURL)
            print('\nGathering of item information Complete!')
            
        elif(choice == 6):
            print('\nStarting key reassignments to beer')
            BeerCategoryTree = reassignKeys(BeerCategoryTree)
            print('\nKey reassignment complete!')
            
        elif(choice == 7):
            
            wordBankChoice = -1
            while wordBankChoice < 0 or wordBankChoice > 5:
                wordBankChoice = -1
                print('\nWord Count Options')
                print('0. Go Back')
                print('1. Compile Word Counts')
                print('2. Save Word Counts to a separate file - NOT YET AVAILABLE')
                print('3. Save combined word count to file')
                print('4. Compile features matrix based on keywords')
                print('5. Save All Current Beer Data as One CSV')
                try:
                    wordBankChoice = int(input())
                except:
                    print('\nInvalid choice, please choose wisely.')
                if wordBankChoice == 1:
                    print('\nCompiling Word Counts, this will take a few minutes..')
                    BeerCategoryTree = compileWordCounts(BeerCategoryTree)
                    print('\nWord Count Compile Complete!')
                elif wordBankChoice == 2:
                    print('\nThis option is not yet available')
                elif wordBankChoice == 3:
                    print('\nSaving Combined Word Count..')
                    combineWordCounts(BeerCategoryTree)
                elif wordBankChoice == 4:
                    print('\nLoading Features Definitions..')
                    beerFeatures = compileFeaturesDefinitions()
                    print('\nConverting Word Counts To Features, this will take a few minutes..')
                    wordCountToFeatures(BeerCategoryTree, beerFeatures)
                    print('\nConverting Word Counts To Features Complete!')
                elif wordBankChoice == 5:
                    print('\nLoading Available Beer Information..')
                    aBeers = loadBeerInformation()
                    print('\nLoading Complete!')
                    print('\nSaving Beer Information to .csv File..')
                    saveBeerInformation(aBeers)
                    print('\nSaving Complete!')
                if wordBankChoice != 0:
                    wordBankChoice = -1
                else:
                    break

        if choice == 8:
            print('\nGetting User Input Information..')
            gatherUserInputInformation()
            print('\nUser Input Information gathered!')

                
        if choice != 0:
            choice = -1
        else:
            break


if __name__ == '__main__':
    main()
